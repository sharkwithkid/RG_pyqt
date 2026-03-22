from engine import (
    inspect_work_root,
    load_all_school_names,
    scan_main_engine,
    run_main_engine,
    scan_diff_engine,
    run_diff_engine,
    get_school_domain,
    get_project_dirs,
    load_notice_templates,
)

from core.roster_log import write_work_result, write_email_sent


import sys
import json
import re
from pathlib import Path

from PyQt6.QtCore import Qt, QDate, QTimer, QThread, QObject, pyqtSignal
from PyQt6.QtGui import QColor

from PyQt6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QMainWindow,
    QPushButton,
    QPlainTextEdit,
    QScrollArea,
    QSplitter,
    QStackedWidget,
    QSpinBox,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QMessageBox,
    QListWidgetItem
)


APP_QSS = """
QMainWindow, QWidget {
    background: #FFFFFF;
    color: #111827;
    font-size: 13px;
}

QLabel#PageTitle {
    font-size: 22px;
    font-weight: 800;
    color: #0F172A;
}

QLabel#PageSubtitle {
    font-size: 14px;
    color: #64748B;
}

QLabel#SectionTitle {
    font-size: 17px;
    font-weight: 700;
    color: #0F172A;
}

QLabel#CardTitle {
    font-size: 14px;
    font-weight: 700;
    color: #0F172A;
}

QLabel#Muted {
    color: #64748B;
    font-size: 12px;
}

QLabel#ValueStrong {
    font-size: 14px;
    font-weight: 800;
    color: #111827;
}

QFrame#Card {
    background: white;
    border: 1px solid #E5E7EB;
    border-radius: 16px;
}

QFrame#SidebarCard {
    background: white;
    border: 1px solid #E5E7EB;
    border-radius: 16px;
}

QFrame#HeaderCard {
    background: white;
    border: 1px solid #E5E7EB;
    border-radius: 18px;
}

QLineEdit, QComboBox, QDateEdit, QTextEdit, QPlainTextEdit, QListWidget, QTableWidget {
    background: #FFFFFF;
    border: 1px solid #D1D5DB;
    border-radius: 10px;
    padding: 8px 10px;
    min-height: 22px;
}

QLineEdit:focus, QComboBox:focus, QDateEdit:focus, QTextEdit:focus, QPlainTextEdit:focus {
    border: 1px solid #2563EB;
}

QPushButton {
    background: #E5E7EB;
    border: none;
    border-radius: 10px;
    padding: 10px 16px;
    font-weight: 700;
    color: #111827;
    min-height: 22px;
}

QPushButton:hover {
    background: #DDE3EA;
}

QPushButton#PrimaryButton {
    background: #2563EB;
    color: white;
    min-height: 22px;
    padding: 12px 18px;
}

QPushButton#PrimaryButton:hover {
    background: #1D4ED8;
}

QPushButton#GhostButton {
    background: transparent;
    border: 1px solid #D1D5DB;
    color: #334155;
}

QPushButton#GhostButton:hover {
    background: #F8FAFC;
}

QTabWidget::pane {
    border: none;
    top: 0px;
}

QTabBar {
    border: none;
}

QTabBar::tab {
    border: none;
}

QTabWidget::pane {
    border: none;
}

QTabBar::tab {
    background: #EAEFF5;
    color: #475569;
    border: none;
    border-top-left-radius: 10px;
    border-top-right-radius: 10px;
    padding: 10px 18px;
    margin-right: 6px;
    font-weight: 700;
}

QTabBar::tab:selected {
    background: white;
    color: #0F172A;
}

QTableWidget {
    gridline-color: #EDF2F7;
    selection-background-color: #DBEAFE;
    selection-color: #111827;
}

QHeaderView::section {
    background: #F8FAFC;
    color: #334155;
    border: none;
    border-bottom: 1px solid #E5E7EB;
    padding: 8px;
    font-weight: 700;
}

QScrollArea {
    border: none;
    background: transparent;
}
"""


def make_card():
    card = QFrame()
    card.setObjectName("Card")
    return card


def make_sidebar_card():
    card = QFrame()
    card.setObjectName("SidebarCard")
    return card


def make_header_card():
    card = QFrame()
    card.setObjectName("HeaderCard")
    return card


CONFIG_PATH = Path("app_config.json")

DEFAULT_APP_CONFIG = {
    "work_root": "",
    "roster_log_path": "",
    "worker_name": "",
    "school_start_date": "",
    "work_date": "",
    "last_arrived_date": "",
    "roster_col_map": {
        "sheet": "",
        "header_row": 0,
        "data_start": 0,
        "col_school": 0,
        "col_email_arr": 0,
        "col_email_snt": 0,
        "col_worker": 0,
        "col_freshmen": 0,
        "col_transfer": 0,
        "col_withdraw": 0,
        "col_teacher": 0,
    },
}


# ── 상태 라벨 공통 스타일 (모듈 레벨 — MainTab/MainWindow 등 여러 클래스에서 공유) ──
_STATUS_STYLE_ERROR = """
    background: #FEE2E2; border: 1px solid #FECACA;
    border-radius: 10px; padding: 8px 12px;
    color: #DC2626; font-weight: 800;
"""
_STATUS_STYLE_WARN = """
    background: #FEF9C3; border: 1px solid #FDE047;
    border-radius: 10px; padding: 8px 12px;
    color: #92400E; font-weight: 800;
"""
_STATUS_STYLE_OK = """
    background: #DCFCE7; border: 1px solid #BBF7D0;
    border-radius: 10px; padding: 8px 12px;
    color: #15803D; font-weight: 800;
"""
_STATUS_STYLE_IDLE = """
    background: #F3F4F6; border: 1px solid #E5E7EB;
    border-radius: 10px; padding: 8px 12px;
    color: #475569; font-weight: 700;
"""
_STATUS_STYLE_RUNNING = """
    background: #DBEAFE; border: 1px solid #BFDBFE;
    border-radius: 10px; padding: 8px 12px;
    color: #1D4ED8; font-weight: 800;
"""


def load_app_config() -> dict:
    if not CONFIG_PATH.exists():
        return DEFAULT_APP_CONFIG.copy()

    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        config = DEFAULT_APP_CONFIG.copy()
        if isinstance(data, dict):
            config.update(data)
        return config
    except Exception:
        return DEFAULT_APP_CONFIG.copy()


def save_app_config(data: dict) -> None:
    config = DEFAULT_APP_CONFIG.copy()
    if isinstance(data, dict):
        config.update(data)

    CONFIG_PATH.write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _work_history_path(school_year: int) -> Path:
    return Path(f"work_history_{school_year}.json")


def load_work_history(school_year: int) -> dict:
    path = _work_history_path(school_year)
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_work_history(school_year: int, school_name: str, entry: dict) -> None:
    history = load_work_history(school_year)
    history[school_name] = entry
    _work_history_path(school_year).write_text(
        json.dumps(history, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _work_history_path(school_year: int) -> Path:
    return Path(f"work_history_{school_year}.json")


def load_work_history(school_year: int) -> dict:
    path = _work_history_path(school_year)
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_work_history(school_year: int, school_name: str, entry: dict) -> None:
    history = load_work_history(school_year)
    history[school_name] = entry
    _work_history_path(school_year).write_text(
        json.dumps(history, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# =========================
# 명단 열 매핑 팝업
# =========================
class RosterColumnMapDialog(QDialog):
    """
    명단 xlsx 파일을 열어 상위 행을 미리보기로 보여주고,
    사용자가 각 역할(학교명, 작업현황 등)의 열을 클릭으로 지정하는 팝업.
    SPSS 변수 지정 팝업과 유사한 UX.
    """

    # 지정할 역할 목록 (key, 표시명, 필수여부)
    ROLES = [
        ("col_school",    "학교명",        True),
        ("col_domain",    "도메인(홈페이지)", True),
        ("col_email_arr", "이메일 도착일자", False),
        ("col_email_snt", "완료 이메일 발송", False),
        ("col_worker",    "작업자",         False),
        ("col_freshmen",  "작업현황(신입생)", False),
        ("col_transfer",  "작업현황(전입생)", False),
        ("col_withdraw",  "작업현황(전출생)", False),
        ("col_teacher",   "작업현황(교직원)", False),
    ]

    ROLE_COLORS = {
        "col_school":    "#DBEAFE",
        "col_domain":    "#E0E7FF",
        "col_email_arr": "#FEF9C3",
        "col_email_snt": "#FEF9C3",
        "col_worker":    "#DCFCE7",
        "col_freshmen":  "#FFE4E6",
        "col_transfer":  "#FFE4E6",
        "col_withdraw":  "#FFE4E6",
        "col_teacher":   "#FFE4E6",
    }

    def __init__(self, xlsx_path: str, existing_map: dict = None, parent=None):
        super().__init__(parent)
        self.xlsx_path = Path(xlsx_path)
        self.result_map = {}          # 최종 결과
        self._col_map = {}            # role_key -> col_index (0-based)
        self._current_role_idx = 0    # 현재 지정 중인 역할 인덱스
        self._sheets = []
        self._headers = []
        self._preview_rows = []
        self._header_row = 0          # 0-based
        self._data_start = 1          # 0-based

        self.setWindowTitle("명단 열 매핑 설정")
        self.setMinimumSize(900, 600)
        self._build_ui()
        self._load_file(existing_map or {})

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)

        # 상단 안내
        guide = QLabel("열 머리글을 클릭해서 각 항목의 열을 지정하세요.")
        guide.setStyleSheet("font-size: 13px; color: #475569;")
        layout.addWidget(guide)

        # 시트 선택
        sheet_row = QHBoxLayout()
        sheet_row.addWidget(QLabel("시트:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.setFixedWidth(200)
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_changed)
        sheet_row.addWidget(self.sheet_combo)

        sheet_row.addSpacing(20)
        sheet_row.addWidget(QLabel("헤더 행:"))
        self.header_spin = QSpinBox()
        self.header_spin.setRange(1, 30)
        self.header_spin.setValue(1)
        self.header_spin.setFixedWidth(60)
        self.header_spin.valueChanged.connect(self._on_header_row_changed)
        sheet_row.addWidget(self.header_spin)
        sheet_row.addStretch()
        layout.addLayout(sheet_row)

        # 역할 선택 패널
        role_panel = QHBoxLayout()
        role_panel.setSpacing(6)

        self.role_buttons = []
        for i, (key, label, required) in enumerate(self.ROLES):
            btn = QPushButton(f"{label}\n-")
            btn.setCheckable(True)
            btn.setFixedHeight(52)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: #F8FAFC;
                    border: 1px solid #CBD5E1;
                    border-radius: 8px;
                    font-size: 11px;
                    padding: 4px 6px;
                    text-align: center;
                }}
                QPushButton:checked {{
                    background: {self.ROLE_COLORS.get(key, '#E0F2FE')};
                    border: 2px solid #3B82F6;
                    font-weight: 700;
                }}
            """)
            btn.clicked.connect(lambda checked, idx=i: self._select_role(idx))
            self.role_buttons.append(btn)
            role_panel.addWidget(btn)

        layout.addLayout(role_panel)

        # 미리보기 테이블
        self.preview_table = QTableWidget()
        self.preview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.preview_table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self.preview_table.horizontalHeader().sectionClicked.connect(self._on_col_header_clicked)
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.preview_table.setMinimumHeight(280)
        layout.addWidget(self.preview_table, 1)

        # 현재 선택 중인 역할 안내
        self.current_role_label = QLabel("")
        self.current_role_label.setStyleSheet(
            "font-size: 13px; font-weight: 700; color: #1D4ED8;"
        )
        layout.addWidget(self.current_role_label)

        # 버튼
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.btn_skip = QPushButton("이 항목 건너뛰기")
        self.btn_skip.setObjectName("GhostButton")
        self.btn_skip.clicked.connect(self._skip_role)
        self.btn_confirm = QPushButton("설정 완료")
        self.btn_confirm.setObjectName("PrimaryButton")
        self.btn_confirm.clicked.connect(self._confirm)
        btn_row.addWidget(self.btn_skip)
        btn_row.addWidget(self.btn_confirm)
        layout.addLayout(btn_row)

    def _load_file(self, existing_map: dict):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(self.xlsx_path), read_only=True, data_only=True)
            self._sheets = wb.sheetnames
            wb.close()
        except Exception as e:
            QMessageBox.critical(self, "파일 오류", f"파일을 열 수 없습니다:\n{e}")
            self.reject()
            return

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.addItems(self._sheets)

        # 기존 설정 복원
        saved_sheet = existing_map.get("sheet", "")
        if saved_sheet in self._sheets:
            self.sheet_combo.setCurrentText(saved_sheet)
        self.sheet_combo.blockSignals(False)

        saved_header = existing_map.get("header_row", 1)
        self.header_spin.setValue(max(1, saved_header))

        # 기존 col_map 복원
        for key, _, _ in self.ROLES:
            v = existing_map.get(key, 0)
            if v:
                self._col_map[key] = v - 1  # 1-based → 0-based

        self._load_sheet_preview()
        self._select_role(0)
        self._refresh_role_buttons()

    def _load_sheet_preview(self):
        sheet_name = self.sheet_combo.currentText()
        header_row = self.header_spin.value()  # 1-based

        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(self.xlsx_path), read_only=True, data_only=True)
            ws = wb[sheet_name]

            rows_data = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows_data.append([str(v) if v is not None else "" for v in row])
                if i >= header_row + 14:  # 헤더 + 15행까지만
                    break
            wb.close()
        except Exception:
            return

        if not rows_data:
            return

        h_idx = header_row - 1  # 0-based
        self._header_row = h_idx
        self._data_start = h_idx + 1

        headers = rows_data[h_idx] if h_idx < len(rows_data) else []
        preview = rows_data[self._data_start:self._data_start + 10]

        self._headers = headers
        self._preview_rows = preview

        self._refresh_table()

    def _refresh_table(self):
        headers = self._headers
        preview = self._preview_rows

        self.preview_table.clear()
        self.preview_table.setColumnCount(len(headers))
        self.preview_table.setRowCount(len(preview))
        self.preview_table.setHorizontalHeaderLabels(
            [f"{i+1}\n{h}" for i, h in enumerate(headers)]
        )

        for r, row in enumerate(preview):
            for c, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                self.preview_table.setItem(r, c, item)

        # 이미 지정된 열에 색상 적용
        self._refresh_col_colors()

    def _refresh_col_colors(self):
        col_count = self.preview_table.columnCount()
        row_count = self.preview_table.rowCount()

        # 역할별 색상 맵 (col_index → color)
        col_color = {}
        for key, _, _ in self.ROLES:
            c = self._col_map.get(key)
            if c is not None and 0 <= c < col_count:
                col_color[c] = self.ROLE_COLORS.get(key, "#E0F2FE")

        for c in range(col_count):
            color = col_color.get(c)
            header_item = self.preview_table.horizontalHeaderItem(c)
            for r in range(row_count):
                item = self.preview_table.item(r, c)
                if item:
                    if color:
                        item.setBackground(QColor(color))
                    else:
                        item.setBackground(QColor("#FFFFFF"))

    def _refresh_role_buttons(self):
        for i, (key, label, _) in enumerate(self.ROLES):
            c = self._col_map.get(key)
            if c is not None and c < len(self._headers):
                col_name = self._headers[c] or f"{c+1}열"
                self.role_buttons[i].setText(f"{label}\n{c+1}열 ({col_name})")
            else:
                self.role_buttons[i].setText(f"{label}\n-")

    def _select_role(self, idx: int):
        self._current_role_idx = idx
        for i, btn in enumerate(self.role_buttons):
            btn.setChecked(i == idx)
        key, label, required = self.ROLES[idx]
        req_text = " (필수)" if required else " (선택)"
        self.current_role_label.setText(
            f"▶  '{label}'{req_text} 열을 표 머리글에서 클릭해 주세요."
        )

    def _on_col_header_clicked(self, col_index: int):
        if self._current_role_idx >= len(self.ROLES):
            return
        key, label, _ = self.ROLES[self._current_role_idx]
        self._col_map[key] = col_index
        self._refresh_role_buttons()
        self._refresh_col_colors()

        # 다음 역할로 자동 이동
        next_idx = self._current_role_idx + 1
        if next_idx < len(self.ROLES):
            self._select_role(next_idx)
        else:
            self.current_role_label.setText("✓  모든 항목 지정 완료. '설정 완료'를 눌러주세요.")

    def _skip_role(self):
        next_idx = self._current_role_idx + 1
        if next_idx < len(self.ROLES):
            self._select_role(next_idx)
        else:
            self.current_role_label.setText("✓  모든 항목 지정 완료. '설정 완료'를 눌러주세요.")

    def _on_sheet_changed(self, _):
        self._col_map.clear()
        self._load_sheet_preview()
        self._refresh_role_buttons()
        self._select_role(0)

    def _on_header_row_changed(self, _):
        self._load_sheet_preview()

    def _confirm(self):
        # 필수 항목 확인
        missing = [
            label for key, label, required in self.ROLES
            if required and key not in self._col_map
        ]
        if missing:
            QMessageBox.warning(
                self, "필수 항목 미지정",
                f"다음 항목은 반드시 지정해야 합니다:\n" + "\n".join(missing)
            )
            return

        self.result_map = {
            "sheet":       self.sheet_combo.currentText(),
            "header_row":  self.header_spin.value(),
            "data_start":  self.header_spin.value() + 1,
        }
        for key, _, _ in self.ROLES:
            c = self._col_map.get(key)
            self.result_map[key] = (c + 1) if c is not None else 0  # 1-based

        self.accept()


class SetupPage(QWidget):
    def __init__(self, on_start_clicked):
        super().__init__()
        self.on_start_clicked = on_start_clicked
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(40, 40, 40, 40)
        page_layout.setSpacing(0)

        container = QWidget()
        container.setMaximumWidth(860)

        content = QVBoxLayout(container)
        content.setContentsMargins(0, 0, 0, 0)
        content.setSpacing(0)

        # ── 타이틀 영역 ──────────────────────────────────────
        title = QLabel("리딩게이트 새학기 반편성")
        title.setObjectName("PageTitle")

        subtitle = QLabel("작업 폴더와 기본 정보를 입력하고 시작하세요.")
        subtitle.setObjectName("PageSubtitle")

        content.addWidget(title)
        content.addSpacing(6)
        content.addWidget(subtitle)

        # ── 타이틀 ↔ 툴 사이 여백 ────────────────────────────
        content.addSpacing(32)

        # ── 최근 작업 요약 카드 3개 ───────────────────────────
        summary_row = QHBoxLayout()
        summary_row.setContentsMargins(0, 0, 0, 0)
        summary_row.setSpacing(12)

        self.stat_last_school = StatMiniCard(
            "마지막 작업 학교",
            "-",
            "이전 작업 없음"
        )
        self.stat_last_date = StatMiniCard(
            "작업일",
            "-",
            "이전 작업 없음"
        )
        self.stat_last_step = StatMiniCard(
            "작업 내용",
            "-",
            "이전 작업 없음"
        )

        summary_row.addWidget(self.stat_last_school)
        summary_row.addWidget(self.stat_last_date)
        summary_row.addWidget(self.stat_last_step)
        content.addLayout(summary_row)

        # ── 요약 ↔ 입력 폼 사이 여백 ─────────────────────────
        content.addSpacing(24)

        # ── 입력 폼 (카드 안에 그룹) ──────────────────────────
        form_card = make_card()
        form_card_layout = QVBoxLayout(form_card)
        form_card_layout.setContentsMargins(24, 22, 24, 22)
        form_card_layout.setSpacing(16)

        # 작업 폴더
        self.work_root_input = QLineEdit()
        self.work_root_input.setPlaceholderText("작업 폴더를 선택하거나 직접 입력하세요")

        self.btn_browse_path = QPushButton("찾아보기")
        self.btn_browse_path.setObjectName("GhostButton")
        self.btn_browse_path.setFixedWidth(110)

        path_row = QHBoxLayout()
        path_row.setContentsMargins(0, 0, 0, 0)
        path_row.setSpacing(10)
        path_row.addWidget(self.work_root_input, 1)
        path_row.addWidget(self.btn_browse_path)

        path_wrap = QWidget()
        path_wrap_layout = QVBoxLayout(path_wrap)
        path_wrap_layout.setContentsMargins(0, 0, 0, 0)
        path_wrap_layout.setSpacing(4)
        path_label = QLabel("작업 폴더")
        path_label.setObjectName("Muted")
        path_wrap_layout.addWidget(path_label)
        path_wrap_layout.addLayout(path_row)

        # 명단 파일
        self.roster_log_input = QLineEdit()
        self.roster_log_input.setPlaceholderText("학교전체명단 .xlsx 파일을 선택하세요")

        self.btn_browse_roster_log = QPushButton("찾아보기")
        self.btn_browse_roster_log.setObjectName("GhostButton")
        self.btn_browse_roster_log.setFixedWidth(110)

        roster_log_row = QHBoxLayout()
        roster_log_row.setContentsMargins(0, 0, 0, 0)
        roster_log_row.setSpacing(10)
        roster_log_row.addWidget(self.roster_log_input, 1)
        roster_log_row.addWidget(self.btn_browse_roster_log)

        roster_log_wrap = QWidget()
        roster_log_wrap_layout = QVBoxLayout(roster_log_wrap)
        roster_log_wrap_layout.setContentsMargins(0, 0, 0, 0)
        roster_log_wrap_layout.setSpacing(4)
        roster_log_label = QLabel("학교 전체 명단 파일 (xlsx)")
        roster_log_label.setObjectName("Muted")
        roster_log_wrap_layout.addWidget(roster_log_label)
        roster_log_wrap_layout.addLayout(roster_log_row)

        self.worker_input = QLineEdit()
        self.worker_input.setPlaceholderText("작업자 이름")

        self.open_date_edit = QDateEdit()
        self.open_date_edit.setCalendarPopup(True)
        self.open_date_edit.setDate(QDate.currentDate())
        self.open_date_edit.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.open_date_edit.wheelEvent = lambda e: None

        self.work_date_edit = QDateEdit()
        self.work_date_edit.setCalendarPopup(True)
        self.work_date_edit.setDate(QDate.currentDate())
        self.work_date_edit.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.work_date_edit.wheelEvent = lambda e: None

        date_row = QGridLayout()
        date_row.setContentsMargins(0, 0, 0, 0)
        date_row.setHorizontalSpacing(14)
        date_row.setVerticalSpacing(8)
        date_row.addWidget(LabeledField("개학일", self.open_date_edit), 0, 0)
        date_row.addWidget(LabeledField("작업일", self.work_date_edit), 0, 1)

        form_card_layout.addWidget(path_wrap)
        form_card_layout.addWidget(roster_log_wrap)
        form_card_layout.addWidget(LabeledField("작업자", self.worker_input))
        form_card_layout.addLayout(date_row)

        content.addWidget(form_card)

        # ── 폼 ↔ 버튼 사이 여백 ──────────────────────────────
        content.addSpacing(28)

        # ── 기본 설정 / 작업 시작 버튼 ───────────────────────
        self.btn_save_defaults = QPushButton("기본 설정 저장")
        self.btn_save_defaults.setObjectName("GhostButton")
        self.btn_save_defaults.setFixedSize(150, 40)

        self.btn_load_defaults = QPushButton("기본 설정 불러오기")
        self.btn_load_defaults.setObjectName("GhostButton")
        self.btn_load_defaults.setFixedSize(170, 40)

        sub_btn_row = QHBoxLayout()
        sub_btn_row.setContentsMargins(0, 0, 0, 0)
        sub_btn_row.setSpacing(10)
        sub_btn_row.addStretch()
        sub_btn_row.addWidget(self.btn_save_defaults)
        sub_btn_row.addWidget(self.btn_load_defaults)
        sub_btn_row.addStretch()
        content.addLayout(sub_btn_row)

        content.addSpacing(14)

        self.btn_start = QPushButton("작업 시작  →")
        self.btn_start.setObjectName("PrimaryButton")
        self.btn_start.setFixedSize(330, 48)
        self.btn_start.setEnabled(False)  # 필수값 입력 전까지 비활성

        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(0, 0, 0, 0)
        btn_row.addStretch()
        btn_row.addWidget(self.btn_start)
        btn_row.addStretch()
        content.addLayout(btn_row)

        page_layout.addStretch(1)
        page_layout.addWidget(container, 0, Qt.AlignmentFlag.AlignHCenter)
        page_layout.addStretch(1)
        scroll.setWidget(page)
        outer.addWidget(scroll)

        self.btn_start.clicked.connect(self.on_start_clicked)
        self.work_root_input.textChanged.connect(self._refresh_start_button)
        self.worker_input.textChanged.connect(self._refresh_start_button)

    def _refresh_start_button(self):
        work_root = self.work_root_input.text().strip()
        worker = self.worker_input.text().strip()
        self.btn_start.setEnabled(bool(work_root) and bool(worker))


class LabeledField(QWidget):
    def __init__(self, label_text: str, field: QWidget):
        super().__init__()
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(3)

        label = QLabel(label_text)
        label.setObjectName("Muted")
        root.addWidget(label)
        root.addWidget(field)


class SchoolSearchLineEdit(QLineEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._key_callback = None

    def set_key_callback(self, cb):
        self._key_callback = cb

    def keyPressEvent(self, event):
        if self._key_callback and self._key_callback(event):
            return
        super().keyPressEvent(event)


class StatMiniCard(QFrame):
    def __init__(self, title: str, value: str, desc: str = ""):
        super().__init__()
        self.setObjectName("Card")
        self.setMinimumHeight(108)
        self.setMaximumHeight(108)

        root = QVBoxLayout(self)
        root.setContentsMargins(14, 12, 14, 12)
        root.setSpacing(2)

        self.lbl_title = QLabel(title)
        self.lbl_title.setObjectName("Muted")

        self.lbl_value = QLabel(value)
        self.lbl_value.setObjectName("ValueStrong")
        self.lbl_value.setWordWrap(True)

        self.lbl_desc = QLabel(desc)
        self.lbl_desc.setObjectName("Muted")
        self.lbl_desc.setWordWrap(True)

        root.addWidget(self.lbl_title)
        root.addStretch()
        root.addWidget(self.lbl_value)
        if desc:
            root.addStretch()
            root.addWidget(self.lbl_desc)

    def set_value(self, value: str):
        self.lbl_value.setText(value)

    def set_desc(self, desc: str):
        self.lbl_desc.setText(desc)
        self.lbl_desc.setVisible(bool(desc))


class StepBar(QWidget):
    """상단 가로 스텝바 — 진행 현황 + 탭 역할 겸용"""

    def __init__(self):
        super().__init__()
        self._callbacks = {}
        self._steps = []
        self._current = -1
        self._states = {i: "idle" for i in range(5)}
        self._build_ui()

    def _build_ui(self):
        self.setFixedHeight(72)
        root = QHBoxLayout(self)
        root.setContentsMargins(20, 0, 20, 0)
        root.setSpacing(0)

        labels = ["기본 설정", "학교 선택", "스캔", "실행·결과", "안내문"]
        self._step_widgets = []

        for i, label in enumerate(labels):
            # 원형 번호 + 텍스트 묶음
            col = QVBoxLayout()
            col.setContentsMargins(0, 0, 0, 0)
            col.setSpacing(4)
            col.setAlignment(Qt.AlignmentFlag.AlignCenter)

            circle = QLabel(str(i + 1))
            circle.setFixedSize(32, 32)
            circle.setAlignment(Qt.AlignmentFlag.AlignCenter)
            circle.setStyleSheet("""
                background: #E5E7EB;
                color: #6B7280;
                border-radius: 16px;
                font-size: 13px;
                font-weight: 700;
            """)

            lbl = QLabel(label)
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet("font-size: 12px; color: #6B7280; font-weight: 500;")

            col.addWidget(circle, 0, Qt.AlignmentFlag.AlignHCenter)
            col.addWidget(lbl, 0, Qt.AlignmentFlag.AlignHCenter)

            step_w = QWidget()
            step_w.setLayout(col)
            step_w.setCursor(Qt.CursorShape.PointingHandCursor)
            step_w.setMinimumWidth(90)

            idx = i
            step_w.mousePressEvent = lambda e, n=idx: self._on_click(n)

            def _make_enter(c, l):
                def _enter(e):
                    if "#2563EB" not in c.styleSheet() and "#15803D" not in c.styleSheet():
                        c.setStyleSheet(
                            "background: #DBEAFE; color: #1D4ED8;"
                            "border-radius: 16px; font-size: 13px; font-weight: 700;"
                        )
                        l.setStyleSheet("font-size: 12px; color: #1D4ED8; font-weight: 600;")
                return _enter

            def _make_leave(n):
                def _leave(e):
                    self._restore_state(n)
                return _leave

            step_w.enterEvent = _make_enter(circle, lbl)
            step_w.leaveEvent = _make_leave(i)

            self._step_widgets.append((step_w, circle, lbl))
            root.addWidget(step_w)

            # 연결선 (마지막 단계 제외)
            if i < len(labels) - 1:
                line = QFrame()
                line.setFrameShape(QFrame.Shape.HLine)
                line.setFixedHeight(2)
                line.setStyleSheet("background: #E5E7EB; border: none;")
                line.setObjectName(f"stepline_{i}")
                self._step_widgets.append((None, line, None))
                root.addWidget(line, 1)

    def _on_click(self, idx: int):
        cb = self._callbacks.get(idx)
        if cb:
            cb()

    def set_callback(self, idx: int, cb):
        self._callbacks[idx] = cb

    def set_state(self, idx: int, state: str):
        """state: 'idle' | 'active' | 'done' | 'warn'"""
        self._states[idx] = state
        self._restore_state(idx)

    def _restore_state(self, idx: int):
        state = self._states.get(idx, "idle")
        real_widgets = [(w, c, l) for (w, c, l) in self._step_widgets if w is not None]
        if idx >= len(real_widgets):
            return
        _, circle, lbl = real_widgets[idx]

        if state == "done":
            circle.setText("✓")
            circle.setStyleSheet("""
                background: #DCFCE7; color: #15803D;
                border-radius: 16px; font-size: 13px; font-weight: 700;
            """)
            lbl.setStyleSheet("font-size: 12px; color: #15803D; font-weight: 500;")
            # 연결선도 초록으로
            self._set_line_color(idx, "#16A34A")
        elif state == "active":
            circle.setText(str(idx + 1))
            circle.setStyleSheet("""
                background: #2563EB; color: white;
                border-radius: 16px; font-size: 13px; font-weight: 700;
            """)
            lbl.setStyleSheet("font-size: 12px; color: #2563EB; font-weight: 700;")
        elif state == "warn":
            circle.setText(str(idx + 1))
            circle.setStyleSheet("""
                background: #FEF3C7; color: #B45309;
                border-radius: 16px; font-size: 13px; font-weight: 700;
            """)
            lbl.setStyleSheet("font-size: 12px; color: #B45309; font-weight: 600;")
        else:  # idle
            circle.setText(str(idx + 1))
            circle.setStyleSheet("""
                background: #E5E7EB; color: #6B7280;
                border-radius: 16px; font-size: 13px; font-weight: 700;
            """)
            lbl.setStyleSheet("font-size: 12px; color: #6B7280; font-weight: 500;")

    def _set_line_color(self, left_step_idx: int, color: str):
        # step_widgets에 step/line 교차로 들어있으므로 line idx = left_step_idx * 2 + 1
        line_list_idx = left_step_idx * 2 + 1
        if line_list_idx < len(self._step_widgets):
            _, line, _ = self._step_widgets[line_list_idx]
            if line:
                line.setStyleSheet(f"background: {color}; border: none;")

    def reset(self):
        for i in range(5):
            self.set_state(i, "idle")
        for i in range(4):
            self._set_line_color(i, "#E5E7EB")


class StepItem(QFrame):
    clicked = None  # will be set per instance

    def __init__(self, number: int, title: str, state: str = "대기"):
        super().__init__()
        self.setObjectName("Card")
        self.number = number
        self.title = title
        self._click_callback = None
        self.setCursor(Qt.CursorShape.PointingHandCursor)

        root = QHBoxLayout(self)
        root.setContentsMargins(14, 12, 14, 12)
        root.setSpacing(12)

        self.badge = QLabel(str(number))
        self.badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.badge.setFixedSize(28, 28)
        self.badge.setStyleSheet("""
            background: #E5E7EB;
            color: #334155;
            border-radius: 14px;
            font-weight: 800;
        """)

        text_wrap = QVBoxLayout()
        text_wrap.setContentsMargins(0, 0, 0, 0)
        text_wrap.setSpacing(2)

        title_label = QLabel(title)
        title_label.setStyleSheet("font-weight: 700; color: #0F172A;")

        self.state_label = QLabel(state)
        self.state_label.setObjectName("Muted")

        text_wrap.addWidget(title_label)
        text_wrap.addWidget(self.state_label)

        root.addWidget(self.badge)
        root.addLayout(text_wrap, 1)

    def set_state(self, state: str):
        self.state_label.setText(state)

        if state == "완료":
            self.badge.setStyleSheet("""
                background: #DCFCE7;
                color: #15803D;
                border-radius: 14px;
                font-weight: 800;
            """)
        elif state == "진행 중":
            self.badge.setStyleSheet("""
                background: #DBEAFE;
                color: #1D4ED8;
                border-radius: 14px;
                font-weight: 800;
            """)
        elif state == "주의":
            self.badge.setStyleSheet("""
                background: #FEF3C7;
                color: #B45309;
                border-radius: 14px;
                font-weight: 800;
            """)
        else:
            self.badge.setStyleSheet("""
                background: #E5E7EB;
                color: #334155;
                border-radius: 14px;
                font-weight: 800;
            """)

    def set_click_callback(self, cb):
        self._click_callback = cb

    def mousePressEvent(self, event):
        if self._click_callback:
            self._click_callback()
        super().mousePressEvent(event)

    def set_selected(self, selected: bool):
        if selected:
            self.setStyleSheet("""
                QFrame#Card {
                    border: 2px solid #2563EB;
                    border-radius: 16px;
                }
            """)
        else:
            self.setStyleSheet("")

    def enterEvent(self, event):
        if not self.styleSheet():
            self.setStyleSheet("QFrame#Card { border: 1px solid #93C5FD; border-radius: 16px; }")
        super().enterEvent(event)

    def leaveEvent(self, event):
        if "93C5FD" in self.styleSheet():
            self.setStyleSheet("")
        super().leaveEvent(event)


class StatusPanel(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(14)

        # 1) 현재 작업 (학교 선택 포함)
        info_card = make_sidebar_card()

        info_layout = QVBoxLayout(info_card)
        info_layout.setContentsMargins(16, 16, 16, 16)
        info_layout.setSpacing(8)

        info_title = QLabel("현재 작업")
        info_title.setObjectName("CardTitle")

        school_wrap = QVBoxLayout()
        school_wrap.setContentsMargins(0, 0, 0, 0)
        school_wrap.setSpacing(8)

        self.school_input = SchoolSearchLineEdit()
        self.school_input.setPlaceholderText("학교명을 입력하세요")
        self.school_input.setFixedHeight(40)

        self.school_result_list = QListWidget()
        self.school_result_list.hide()
        self.school_result_list.setMaximumHeight(180)
        self.school_result_list.setSpacing(2)
        self.school_result_list.setVerticalScrollMode(QListWidget.ScrollMode.ScrollPerPixel)

        self.btn_select_school = QPushButton("적용")
        self.btn_select_school.setObjectName("PrimaryButton")
        self.btn_select_school.setFixedWidth(100)
        self.btn_select_school.setFixedHeight(40)

        input_row = QHBoxLayout()
        input_row.setContentsMargins(0, 0, 0, 0)
        input_row.setSpacing(8)
        input_row.addWidget(self.school_input, 1)
        input_row.addWidget(self.btn_select_school)

        school_wrap.addLayout(input_row)
        school_wrap.addWidget(self.school_result_list)

        self.current_school = QLabel("-")
        self.current_school.setStyleSheet("font-size: 17px; font-weight: 800; color: #0F172A;")
        self.current_school.setWordWrap(True)

        self.school_history_label = QLabel("")
        self.school_history_label.setObjectName("Muted")
        self.school_history_label.setWordWrap(True)
        self.school_history_label.hide()

        self.school_status_label = QLabel("학교명을 입력해 검색하세요.")
        self.school_status_label.setObjectName("Muted")
        self.school_status_label.setWordWrap(True)

        # 작업 이력 라벨
        self.last_work_label = QLabel("")
        self.last_work_label.setObjectName("Muted")
        self.last_work_label.setWordWrap(True)
        self.last_work_label.hide()

        self.latest_log_label = QLabel("")
        self.latest_log_label.setObjectName("Muted")
        self.latest_log_label.setWordWrap(True)
        self.latest_log_label.hide()
        self.latest_log_label.setMaximumHeight(0)

        # 이메일 도착일자
        email_date_label = QLabel("이메일 도착일자")
        email_date_label.setObjectName("Muted")
        self.email_arrived_date = QDateEdit()
        self.email_arrived_date.setCalendarPopup(True)
        self.email_arrived_date.setDate(QDate.currentDate())
        self.email_arrived_date.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.email_arrived_date.wheelEvent = lambda e: None
        self.email_arrived_date.setFixedHeight(36)

        self.chk_email_arrived = QCheckBox("도착일 기록")
        self.chk_email_arrived.setChecked(False)
        self.email_arrived_date.setEnabled(False)
        self.chk_email_arrived.toggled.connect(self.email_arrived_date.setEnabled)

        email_row = QHBoxLayout()
        email_row.setContentsMargins(0, 0, 0, 0)
        email_row.setSpacing(8)
        email_row.addWidget(self.chk_email_arrived)
        email_row.addWidget(self.email_arrived_date, 1)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color: #E5E7EB;")

        self.current_worker = QLabel("-")
        self.current_worker.setObjectName("Muted")

        self.current_work_date = QLabel("-")
        self.current_work_date.setObjectName("Muted")

        # 학년별 기준 학년도 토글
        self.btn_grade_map = QPushButton("학년도 아이디 규칙 보기")
        self.btn_grade_map.setObjectName("GhostButton")
        self.btn_grade_map.setCheckable(True)

        self.grade_map_widget = QWidget()
        self.grade_map_widget.setVisible(False)

        grade_grid = QGridLayout(self.grade_map_widget)
        grade_grid.setContentsMargins(4, 4, 4, 4)
        grade_grid.setSpacing(6)

        grade_hint = QLabel("스캔 후 자동 감지 · 직접 수정 가능")
        grade_hint.setObjectName("Muted")
        grade_hint.setStyleSheet("font-size: 11px;")
        grade_grid.addWidget(grade_hint, 0, 0, 1, 2)

        self.grade_spinboxes = {}
        self.grade_row_widgets = {}  # {grade: (label, spinbox)}
        for i in range(6):
            grade = i + 1
            lbl_g = QLabel(f"{grade}학년")
            lbl_g.setObjectName("Muted")

            spin = QSpinBox()
            spin.setRange(0, 2099)
            spin.setSpecialValueText("-")
            spin.setValue(0)
            spin.setFixedWidth(80)
            spin.setFixedHeight(28)
            spin.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
            spin.wheelEvent = lambda e: None
            spin.setStyleSheet("""
                QSpinBox {
                    border: 1px solid #D1D5DB;
                    border-radius: 6px;
                    padding: 2px 6px;
                    font-size: 12px;
                    font-weight: 700;
                    color: #0F172A;
                }
            """)

            grade_grid.addWidget(lbl_g, i + 1, 0)
            grade_grid.addWidget(spin, i + 1, 1)
            self.grade_spinboxes[grade] = spin
            self.grade_row_widgets[grade] = (lbl_g, spin)

        self.btn_grade_map_apply = QPushButton("적용")
        self.btn_grade_map_apply.setObjectName("PrimaryButton")
        self.btn_grade_map_apply.setFixedHeight(30)
        self.btn_grade_map_apply.setStyleSheet("""
            QPushButton {
                background: #2563EB;
                color: white;
                border-radius: 8px;
                padding: 4px 14px;
                font-weight: 700;
                font-size: 12px;
            }
            QPushButton:hover { background: #1D4ED8; }
        """)
        grade_grid.addWidget(self.btn_grade_map_apply, 7, 0, 1, 2)

        # 하위 호환 — update_grade_map이 참조하는 grade_labels 유지
        self.grade_labels = {g: spin for g, spin in self.grade_spinboxes.items()}

        def _toggle_grade_map():
            visible = self.btn_grade_map.isChecked()
            self.grade_map_widget.setVisible(visible)
            self.btn_grade_map.setText(
                "학년도 아이디 규칙 숨기기" if visible else "학년도 아이디 규칙 보기"
            )

        self.btn_grade_map.clicked.connect(_toggle_grade_map)

        info_layout.addWidget(info_title)
        info_layout.addLayout(school_wrap)
        info_layout.addWidget(self.current_school)
        info_layout.addWidget(self.school_history_label)
        info_layout.addWidget(self.school_status_label)
        info_layout.addWidget(self.last_work_label)
        info_layout.addWidget(self.latest_log_label)

        # 작업자/작업일 하위 호환용 더미 (상단 헤더에 있으므로 사이드바에서 제거)
        self.current_worker = QLabel("")
        self.current_worker.setMaximumHeight(0)
        self.current_worker.hide()
        self.current_work_date = QLabel("")
        self.current_work_date.setMaximumHeight(0)
        self.current_work_date.hide()

        sep1 = QFrame()
        sep1.setFrameShape(QFrame.Shape.HLine)
        sep1.setStyleSheet("color: #E5E7EB;")
        info_layout.addWidget(sep1)

        # 학년도 아이디 규칙
        self.btn_grade_map = QPushButton("학년도 아이디 규칙 보기")
        self.btn_grade_map.setObjectName("GhostButton")
        self.btn_grade_map.setCheckable(True)

        self.grade_map_widget = QWidget()
        self.grade_map_widget.setVisible(False)

        grade_grid = QGridLayout(self.grade_map_widget)
        grade_grid.setContentsMargins(4, 4, 4, 4)
        grade_grid.setSpacing(6)

        grade_hint = QLabel("스캔 후 자동 감지 · 직접 수정 가능")
        grade_hint.setObjectName("Muted")
        grade_hint.setStyleSheet("font-size: 11px;")
        grade_grid.addWidget(grade_hint, 0, 0, 1, 2)

        self.grade_spinboxes = {}
        self.grade_row_widgets = {}
        for i in range(6):
            grade = i + 1
            lbl_g = QLabel(f"{grade}학년")
            lbl_g.setObjectName("Muted")
            spin = QSpinBox()
            spin.setRange(0, 2099)
            spin.setSpecialValueText("-")
            spin.setValue(0)
            spin.setFixedWidth(80)
            spin.setFixedHeight(28)
            spin.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
            spin.wheelEvent = lambda e: None
            spin.setStyleSheet("""
                QSpinBox {
                    border: 1px solid #D1D5DB;
                    border-radius: 6px;
                    padding: 2px 6px;
                    font-size: 12px;
                    font-weight: 700;
                    color: #0F172A;
                }
            """)
            grade_grid.addWidget(lbl_g, i + 1, 0)
            grade_grid.addWidget(spin, i + 1, 1)
            self.grade_spinboxes[grade] = spin
            self.grade_row_widgets[grade] = (lbl_g, spin)

        self.btn_grade_map_apply = QPushButton("적용")
        self.btn_grade_map_apply.setObjectName("PrimaryButton")
        self.btn_grade_map_apply.setFixedHeight(30)
        self.btn_grade_map_apply.setStyleSheet("""
            QPushButton {
                background: #2563EB; color: white;
                border-radius: 8px; padding: 4px 14px;
                font-weight: 700; font-size: 12px;
            }
            QPushButton:hover { background: #1D4ED8; }
        """)
        grade_grid.addWidget(self.btn_grade_map_apply, 7, 0, 1, 2)

        self.grade_labels = {g: spin for g, spin in self.grade_spinboxes.items()}

        def _toggle_grade_map():
            visible = self.btn_grade_map.isChecked()
            self.grade_map_widget.setVisible(visible)
            self.btn_grade_map.setText(
                "학년도 아이디 규칙 숨기기" if visible else "학년도 아이디 규칙 보기"
            )
        self.btn_grade_map.clicked.connect(_toggle_grade_map)

        info_layout.addWidget(self.btn_grade_map)
        info_layout.addWidget(self.grade_map_widget)

        sep2 = QFrame()
        sep2.setFrameShape(QFrame.Shape.HLine)
        sep2.setStyleSheet("color: #E5E7EB;")
        info_layout.addWidget(sep2)

        # 도착일 — 체크박스로 기록 확정, 날짜는 항상 수정 가능
        self.chk_email_arrived = QCheckBox("도착일 기록")
        self.chk_email_arrived.setChecked(False)
        self.email_arrived_date = QDateEdit()
        self.email_arrived_date.setCalendarPopup(True)
        self.email_arrived_date.setDate(QDate.currentDate())
        self.email_arrived_date.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.email_arrived_date.wheelEvent = lambda e: None
        self.email_arrived_date.setFixedHeight(32)

        arrived_row = QHBoxLayout()
        arrived_row.setContentsMargins(0, 0, 0, 0)
        arrived_row.setSpacing(8)
        arrived_row.addWidget(self.chk_email_arrived)
        arrived_row.addWidget(self.email_arrived_date, 1)
        info_layout.addLayout(arrived_row)

        # 발송일 — 체크박스로 기록 확정, 날짜는 항상 수정 가능
        self.chk_email_sent = QCheckBox("발송일 기록")
        self.chk_email_sent.setChecked(False)
        self.email_sent_date = QDateEdit()
        self.email_sent_date.setCalendarPopup(True)
        self.email_sent_date.setDate(QDate.currentDate())
        self.email_sent_date.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.email_sent_date.wheelEvent = lambda e: None
        self.email_sent_date.setFixedHeight(32)

        sent_row = QHBoxLayout()
        sent_row.setContentsMargins(0, 0, 0, 0)
        sent_row.setSpacing(8)
        sent_row.addWidget(self.chk_email_sent)
        sent_row.addWidget(self.email_sent_date, 1)
        info_layout.addLayout(sent_row)

        # 명단 기록
        self.btn_record_roster_sidebar = QPushButton("명단 기록")
        self.btn_record_roster_sidebar.setObjectName("GhostButton")
        self.btn_record_roster_sidebar.setEnabled(False)
        info_layout.addWidget(self.btn_record_roster_sidebar)

        sep_bot = QFrame()
        sep_bot.setFrameShape(QFrame.Shape.HLine)
        sep_bot.setStyleSheet("color: #E5E7EB;")
        info_layout.addWidget(sep_bot)

        # 새 학교 시작
        self.btn_new_school = QPushButton("새 학교 시작 →")
        self.btn_new_school.setStyleSheet("""
            QPushButton {
                background: #F1F5F9; border: 1px solid #E2E8F0;
                border-radius: 10px; padding: 10px 16px;
                font-weight: 700; color: #334155;
            }
            QPushButton:hover { background: #E2E8F0; }
        """)
        info_layout.addWidget(self.btn_new_school)

        root.addWidget(info_card)
        root.addStretch()

        self.progress_card_widget = QWidget()
        self._step_tab_map = {}
        self._tab_switch_callback = None
        self._setup_callback = None
        self._school_callback = None

        self.setFixedWidth(290)

    def update_grade_map(self, mapping: dict = None, state: str = "default"):
        """
        state:
          "default"    → 스캔 전 (0으로 리셋)
          "not_needed" → 이번 작업은 학생명부 불필요
          "no_roster"  → 명부 필요하나 없거나 읽지 못함 (0 유지, 수동 입력 유도)
          "ok"         → 값 표시 (mapping 사용)
        SpinBox 값 0 = "-" (specialValueText)
        """
        if state in ("default", "not_needed"):
            for spin in self.grade_spinboxes.values():
                spin.setValue(0)
                spin.setEnabled(state == "default")
            return

        if state == "no_roster":
            # 명부 없음 → 수동 입력 가능하도록 활성화, 값은 0 유지
            for spin in self.grade_spinboxes.values():
                spin.setValue(0)
                spin.setEnabled(True)
            return

        # state == "ok"
        for grade, spin in self.grade_spinboxes.items():
            spin.setEnabled(True)
            val = mapping.get(grade) if mapping else None
            spin.setValue(int(val) if val is not None else 0)

    def get_grade_year_overrides(self) -> dict:
        """사용자가 설정한 학년별 학년도 값 반환 {grade: year}. 0이면 제외."""
        result = {}
        for grade, spin in self.grade_spinboxes.items():
            v = spin.value()
            if v > 0:
                result[grade] = v
        return result

    def set_grade_count(self, school_name: str):
        """학교명 기준으로 표시할 학년 수 결정 (초등 6, 중·고등 3)."""
        name = (school_name or "").strip()
        # 마지막 글자로 학교 종류 판단
        last = name[-1] if name else ""
        if last in ("중", "고"):
            max_grade = 3
        else:
            max_grade = 6  # 초등 or 기타

        for grade, (lbl, spin) in self.grade_row_widgets.items():
            visible = grade <= max_grade
            lbl.setVisible(visible)
            spin.setVisible(visible)
            if not visible:
                spin.setValue(0)

    def set_tab_switch_callback(self, cb):
        self._tab_switch_callback = cb

    def set_setup_callback(self, cb):
        self._setup_callback = cb

    def set_school_callback(self, cb):
        self._school_callback = cb

    def _on_setup_clicked(self):
        if self._setup_callback:
            self._setup_callback()

    def _on_school_clicked(self):
        if self._school_callback:
            self._school_callback()

    def _on_step_clicked(self, tab_index: int):
        if self._tab_switch_callback:
            self._tab_switch_callback(tab_index)
        self.on_tab_changed(tab_index)

    def on_tab_changed(self, tab_index: int):
        for idx, step in self._step_tab_map.items():
            step.set_selected(idx == tab_index)


class ScanWorker(QObject):
    finished = pyqtSignal(object)
    failed = pyqtSignal(str)

    def __init__(self, work_root, school_name, school_start_date, work_date,
                 roster_xlsx=None, col_map=None):
        super().__init__()
        self.work_root = work_root
        self.school_name = school_name
        self.school_start_date = school_start_date
        self.work_date = work_date
        self.roster_xlsx = roster_xlsx
        self.col_map = col_map

    def run(self):
        try:
            result = scan_main_engine(
                work_root=self.work_root,
                school_name=self.school_name,
                school_start_date=self.school_start_date,
                work_date=self.work_date,
                roster_xlsx=self.roster_xlsx,
                col_map=self.col_map,
            )
            self.finished.emit(result)
        except Exception as e:
            import traceback
            self.failed.emit(traceback.format_exc())


class RunWorker(QObject):
    finished = pyqtSignal(object)
    failed = pyqtSignal(str)

    def __init__(self, scan, work_date, school_start_date, layout_overrides=None,
                 school_kind_override=None):
        super().__init__()
        self.scan = scan
        self.work_date = work_date
        self.school_start_date = school_start_date
        self.layout_overrides = layout_overrides or {}
        self.school_kind_override = school_kind_override

    def run(self):
        try:
            result = run_main_engine(
                scan=self.scan,
                work_date=self.work_date,
                school_start_date=self.school_start_date,
                layout_overrides=self.layout_overrides,
                school_kind_override=self.school_kind_override,
            )
            self.finished.emit(result)
        except Exception as e:
            import traceback
            self.failed.emit(traceback.format_exc())


class AppHeader(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        card = make_header_card()
        root_outer = QVBoxLayout(self)
        root_outer.setContentsMargins(0, 0, 0, 0)
        root_outer.addWidget(card)

        root = QHBoxLayout(card)
        root.setContentsMargins(22, 18, 22, 18)
        root.setSpacing(18)

        left = QVBoxLayout()
        left.setContentsMargins(0, 0, 0, 0)
        left.setSpacing(4)

        app_name = QLabel("리딩게이트 반이동 자동화")
        app_name.setStyleSheet("font-size: 21px; font-weight: 800; color: #0F172A;")

        self.school_name = QLabel("학교를 선택해 주세요")
        self.school_name.setObjectName("Muted")

        left.addWidget(app_name)
        left.addWidget(self.school_name)

        # 모드 전환 버튼
        mode_row = QHBoxLayout()
        mode_row.setSpacing(6)

        self.btn_mode_main = QPushButton("새학기 반편성")
        self.btn_mode_diff = QPushButton("재학생 명단 비교")

        for btn in [self.btn_mode_main, self.btn_mode_diff]:
            btn.setCheckable(True)
            btn.setStyleSheet("""
                QPushButton {
                    background: #F1F5F9;
                    border: 1px solid #E2E8F0;
                    border-radius: 10px;
                    padding: 8px 16px;
                    font-weight: 700;
                    color: #64748B;
                }
                QPushButton:checked {
                    background: #2563EB;
                    border: 1px solid #2563EB;
                    color: white;
                }
            """)
            mode_row.addWidget(btn)

        self.btn_mode_main.setChecked(True)

        right = QHBoxLayout()
        right.setSpacing(10)

        self.badge_task = QLabel("작업 -")
        self.badge_worker = QLabel("작업자 -")
        self.badge_status = QLabel("상태 대기")

        for badge in [self.badge_task, self.badge_worker, self.badge_status]:
            badge.setStyleSheet("""
                background: #F8FAFC;
                border: 1px solid #E5E7EB;
                border-radius: 10px;
                padding: 8px 12px;
                color: #334155;
                font-weight: 700;
            """)
            right.addWidget(badge)

        root.addLayout(left)
        root.addLayout(mode_row)
        root.addStretch(1)
        root.addLayout(right)


class SchoolHeader(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(14)

        top_card = make_card()
        top_layout = QVBoxLayout(top_card)
        top_layout.setContentsMargins(18, 18, 18, 18)
        top_layout.setSpacing(12)

        title = QLabel("학교 선택")
        title.setObjectName("CardTitle")

        desc = QLabel("작업할 학교를 선택하고 현재 입력 파일 상태를 확인합니다.")
        desc.setObjectName("Muted")

        row = QHBoxLayout()
        row.setSpacing(10)

        self.school_combo = QComboBox()
        self.school_combo.setEditable(True)
        self.school_combo.setCurrentIndex(-1)

        self.btn_select_school = QPushButton("학교 적용")
        self.btn_select_school.setObjectName("PrimaryButton")
        self.btn_select_school.setFixedWidth(120)

        row.addWidget(self.school_combo, 1)
        row.addWidget(self.btn_select_school)

        self.latest_log_label = QLabel("이전 작업 이력 없음")
        self.latest_log_label.setObjectName("Muted")

        top_layout.addWidget(title)
        top_layout.addWidget(desc)
        top_layout.addLayout(row)
        top_layout.addWidget(self.latest_log_label)

        file_card = make_card()
        file_layout = QVBoxLayout(file_card)
        file_layout.setContentsMargins(18, 18, 18, 18)
        file_layout.setSpacing(10)

        file_title = QLabel("입력 파일 상태")
        file_title.setObjectName("CardTitle")

        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(10)

        self.file_status_labels = []

        names = [
            "학생명부.xlsx",
            "신입생 명단.xlsx",
            "전입생 명단.xlsx",
            "전출생 명단.xlsx",
            "교사 명단.xlsx",
        ]

        for i, text in enumerate(names):
            pill = QLabel(f"미확인 · {text}")
            pill.setStyleSheet("""
                background: #F8FAFC;
                border: 1px solid #E5E7EB;
                border-radius: 10px;
                padding: 9px 12px;
                color: #334155;
                font-weight: 600;
            """)
            grid.addWidget(pill, i // 2, i % 2)
            self.file_status_labels.append(pill)

        file_layout.addWidget(file_title)
        file_layout.addLayout(grid)

        root.addWidget(top_card)
        root.addWidget(file_card)


class SchoolPage(QWidget):
    def __init__(self, on_confirmed):
        super().__init__()
        self.on_confirmed = on_confirmed
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(40, 0, 40, 0)
        outer.setSpacing(0)

        container = QWidget()
        container.setMaximumWidth(860)

        content = QVBoxLayout(container)
        content.setContentsMargins(0, 0, 0, 0)
        content.setSpacing(0)

        # 타이틀
        title = QLabel("학교 선택")
        title.setObjectName("PageTitle")
        subtitle = QLabel("작업할 학교를 선택하고 입력 파일 상태를 확인하세요.")
        subtitle.setObjectName("PageSubtitle")

        content.addWidget(title)
        content.addSpacing(6)
        content.addWidget(subtitle)
        content.addSpacing(32)

        # 학교 선택 카드
        school_card = make_card()
        school_layout = QVBoxLayout(school_card)
        school_layout.setContentsMargins(24, 22, 24, 22)
        school_layout.setSpacing(14)

        school_label = QLabel("학교")
        school_label.setObjectName("Muted")

        combo_row = QHBoxLayout()
        combo_row.setSpacing(10)

        self.school_combo = QComboBox()
        self.school_combo.setEditable(True)
        self.school_combo.setCurrentIndex(-1)
        self.school_combo.setPlaceholderText("학교명을 입력하거나 선택하세요")

        self.btn_select_school = QPushButton("학교 적용")
        self.btn_select_school.setObjectName("PrimaryButton")
        self.btn_select_school.setFixedWidth(120)

        combo_row.addWidget(self.school_combo, 1)
        combo_row.addWidget(self.btn_select_school)

        self.latest_log_label = QLabel("이전 작업 이력 없음")
        self.latest_log_label.setObjectName("Muted")

        school_layout.addWidget(school_label)
        school_layout.addLayout(combo_row)
        school_layout.addWidget(self.latest_log_label)

        content.addWidget(school_card)
        content.addSpacing(20)

        # 입력 파일 상태 카드
        file_card = make_card()
        file_layout = QVBoxLayout(file_card)
        file_layout.setContentsMargins(24, 22, 24, 22)
        file_layout.setSpacing(14)

        file_title = QLabel("입력 파일 상태")
        file_title.setObjectName("SectionTitle")

        file_desc = QLabel("학교 폴더 안의 파일이 올바르게 준비되어 있는지 확인하세요.")
        file_desc.setObjectName("Muted")

        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(10)

        self.file_status_labels = []
        names = [
            "학생명부.xlsx",
            "신입생 명단.xlsx",
            "전입생 명단.xlsx",
            "전출생 명단.xlsx",
            "교사 명단.xlsx",
        ]
        for i, text in enumerate(names):
            pill = QLabel(f"미확인 · {text}")
            pill.setStyleSheet("""
                background: #F8FAFC;
                border: 1px solid #E5E7EB;
                border-radius: 10px;
                padding: 10px 14px;
                color: #334155;
                font-weight: 600;
            """)
            grid.addWidget(pill, i // 2, i % 2)
            self.file_status_labels.append(pill)

        file_layout.addWidget(file_title)
        file_layout.addWidget(file_desc)
        file_layout.addLayout(grid)

        content.addWidget(file_card)
        content.addSpacing(28)

        # 작업 시작 버튼
        self.btn_confirm = QPushButton("작업 시작  →")
        self.btn_confirm.setObjectName("PrimaryButton")
        self.btn_confirm.setFixedHeight(48)

        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(0, 0, 0, 0)
        btn_row.addStretch()
        btn_row.addWidget(self.btn_confirm)
        content.addLayout(btn_row)

        outer.addStretch(1)
        outer.addWidget(container, 0, Qt.AlignmentFlag.AlignHCenter)
        outer.addStretch(1)

        self.btn_confirm.clicked.connect(self.on_confirmed)


class MainTab(QWidget):

    _open_scan_file_signal = pyqtSignal(str)
    _open_run_file_signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.scan_preview_data = {}
        self.current_preview_kind = None

        self.run_preview_data = {}
        self.current_run_file = None
        self.run_sheet_tables = {}
        self.run_sheet_raw_rows = {}

        self._build_ui()

    def populate_run_sheet_tabs(self):
        self.run_sheet_tabs.clear()
        self.run_sheet_tables = {}
        self.run_sheet_raw_rows = {}

        if not self.current_run_file:
            return

        data = self.run_preview_data.get(self.current_run_file, {})
        sheets = data.get("sheets", {})

        for sheet_name, sheet_data in sheets.items():
            headers = sheet_data.get("headers", [])
            rows = sheet_data.get("rows", [])

            table = QTableWidget()
            table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            table.horizontalHeader().setStretchLastSection(False)

            self.run_sheet_tables[sheet_name] = table
            self.run_sheet_raw_rows[sheet_name] = {
                "headers": headers,
                "rows": rows,
            }

            tab_page = QWidget()
            tab_layout = QVBoxLayout(tab_page)
            tab_layout.setContentsMargins(0, 0, 0, 0)
            tab_layout.addWidget(table)

            self.run_sheet_tabs.addTab(tab_page, sheet_name)

        self.filter_current_run_sheet()
    

    def filter_current_run_sheet(self):
        current_index = self.run_sheet_tabs.currentIndex()
        if current_index < 0:
            self.run_preview_info.setText("시트: - | 행 수: -")
            return

        sheet_name = self.run_sheet_tabs.tabText(current_index)
        table = self.run_sheet_tables.get(sheet_name)
        raw = self.run_sheet_raw_rows.get(sheet_name, {})

        if not table:
            return

        headers = raw.get("headers", [])
        rows = raw.get("rows", [])

        keyword = self.run_view_search.text().strip().lower()
        dup_only = self.btn_run_dup_only.isChecked()

        # 이름/학년 컬럼 탐색
        NAME_KEYWORDS = ["이름", "성명", "학생이름"]
        name_col = next(
            (i for kw in NAME_KEYWORDS for i, h in enumerate(headers) if kw in str(h)),
            None
        )
        grade_col = next(
            (i for i, h in enumerate(headers) if "학년" in str(h)),
            None
        )

        # 동명이인 인덱스 사전 계산
        # suffix(A/B/AA...) 제거한 원본 이름 기준으로 판정
        dup_indices: set = set()
        if dup_only and name_col is not None:
            import re as _re
            from collections import Counter
            def _base_name(nm: str) -> str:
                # 한글이 포함된 경우에만 끝 알파벳 suffix 제거
                if _re.search(r"[\uAC00-\uD7A3]", nm):
                    return _re.sub(r"[A-Z]+$", "", nm).strip()
                return nm
            key_count: Counter = Counter()
            for row in rows:
                nm = _base_name(str(row[name_col]).strip()) if name_col < len(row) else ""
                grade = str(row[grade_col]).strip() if (grade_col is not None and grade_col < len(row)) else ""
                if nm:
                    key_count[(grade, nm)] += 1
            for idx, row in enumerate(rows):
                nm = _base_name(str(row[name_col]).strip()) if name_col < len(row) else ""
                grade = str(row[grade_col]).strip() if (grade_col is not None and grade_col < len(row)) else ""
                if nm and key_count[(grade, nm)] >= 2:
                    dup_indices.add(idx)

        filtered_rows = []
        for idx, row in enumerate(rows):
            row_values = [str(v) for v in row]
            if keyword:
                if keyword not in " ".join(row_values).lower():
                    continue
            if dup_only and idx not in dup_indices:
                continue
            filtered_rows.append((idx, row))

        table.clear()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(filtered_rows))
        # 열 너비 균등화 (7p-15)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setStretchLastSection(False)

        for r, (idx, row) in enumerate(filtered_rows):
            for c, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                # 동명이인 노란색 강조
                if idx in dup_indices:
                    item.setBackground(QColor("#FEF9C3"))
                table.setItem(r, c, item)

        table.resizeColumnsToContents()
        self.run_preview_info.setText(f"시트: {sheet_name} | 행 수: {len(filtered_rows)}")


    def update_run_preview_info(self):
        self.filter_current_run_sheet()


    def fit_scan_table_height(self, visible_rows=4, row_h=28):
        row_count = min(visible_rows, self.scan_table.rowCount())
        if row_count <= 0:
            return

        for i in range(self.scan_table.rowCount()):
            self.scan_table.setRowHeight(i, row_h)

        header_h = self.scan_table.horizontalHeader().height()
        frame_h = self.scan_table.frameWidth() * 2

        hbar_h = 0
        if self.scan_table.horizontalScrollBar().isVisible():
            hbar_h = self.scan_table.horizontalScrollBar().height()

        extra = 4
        total_h = header_h + (row_h * row_count) + frame_h + hbar_h + extra

        self.scan_table.setFixedHeight(total_h)


    def _make_scroll(self, content_widget: QWidget) -> QScrollArea:
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setWidget(content_widget)
        return scroll
    

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        self.run_preview_data = {}
        self.current_run_file = None

        # QTabWidget 대신 QStackedWidget — 사이드바 단계 클릭으로 전환
        self.tabs = QStackedWidget()
        self.tabs.addWidget(self._build_scan_tab())   # index 0
        self.tabs.addWidget(self._build_run_tab())    # index 1
        self.tabs.addWidget(self._build_notice_tab()) # index 2
        outer.addWidget(self.tabs, 1)

    def show_step(self, idx: int):
        """사이드바 단계 클릭 시 호출 — 해당 인덱스 화면으로 전환."""
        self.tabs.setCurrentIndex(idx)

    def _build_scan_tab(self):
        page = QWidget()
        root = QVBoxLayout(page)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        wrap = QWidget()
        wrap_layout = QVBoxLayout(wrap)
        wrap_layout.setContentsMargins(0, 0, 0, 0)
        wrap_layout.setSpacing(0)
        scan_box = self._build_scan_box()
        wrap_layout.addWidget(scan_box)
        wrap_layout.addStretch()
        scroll.setWidget(wrap)

        root.addWidget(scroll)
        return page
    
    def _build_run_tab(self):
        page = QWidget()
        root = QVBoxLayout(page)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(0)
        root.addWidget(self._build_run_box(), 1)
        return page

    def _build_notice_tab(self):
        page = QWidget()
        root = QVBoxLayout(page)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(0)
        root.addWidget(self._build_notice_box(), 1)
        return page

    def _build_diff_tab(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)

        content = QWidget()
        root = QVBoxLayout(content)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(14)

        root.addWidget(self._build_diff_run_box())
        root.addWidget(self._build_diff_result_box())
        root.addStretch()

        layout.addWidget(self._make_scroll(content))
        return page

    def _build_scan_box(self):
        card = make_card()
        from PyQt6.QtWidgets import QSizePolicy
        card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
        outer = QVBoxLayout(card)
        outer.setContentsMargins(14, 10, 14, 10)
        outer.setSpacing(6)

        head_row = QHBoxLayout()
        head_left = QVBoxLayout()
        head_left.setContentsMargins(0, 0, 0, 0)
        head_left.setSpacing(4)

        title = QLabel("스캔 검수")
        title.setObjectName("CardTitle")

        desc = QLabel("입력 파일 구조를 먼저 확인하고, 의심 행을 전체 미리보기에서 검수합니다.")
        desc.setObjectName("Muted")

        head_left.addWidget(title)
        head_left.addWidget(desc)

        action_row = QHBoxLayout()
        action_row.setSpacing(8)

        self.btn_scan = QPushButton("파일 내용 스캔")
        self.btn_scan.setObjectName("PrimaryButton")
        self.btn_show_scan_log = QPushButton("스캔 로그 보기")
        self.btn_show_scan_log.setObjectName("GhostButton")

        self.scan_status_label = QLabel("스캔 전")
        self.scan_status_label.setStyleSheet(_STATUS_STYLE_IDLE)
        self.scan_status_label.setFixedWidth(90)
        self.scan_status_label.setFixedHeight(40)
        self.scan_status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.btn_goto_run_tab = QPushButton("→")
        self.btn_goto_run_tab.setObjectName("PrimaryButton")
        self.btn_goto_run_tab.setEnabled(False)
        self.btn_goto_run_tab.setFixedSize(40, 40)
        self.btn_goto_run_tab.setStyleSheet("""
            QPushButton {
                background: #2563EB; color: white;
                border-radius: 10px; font-size: 18px; font-weight: 700;
                padding: 0px;
            }
            QPushButton:hover { background: #1D4ED8; }
            QPushButton:disabled { background: #93C5FD; }
        """)
        self.btn_goto_run_tab.setVisible(False)

        action_row.addStretch()
        action_row.addWidget(self.btn_scan)
        action_row.addWidget(self.btn_show_scan_log)
        action_row.addWidget(self.scan_status_label)
        action_row.addWidget(self.btn_goto_run_tab)

        head_row.addLayout(head_left, 1)
        head_row.addLayout(action_row)

        self.scan_message = QLabel("파일 내용 스캔을 실행해 주세요.")
        self.scan_message.setObjectName("Muted")
        self.scan_message.setWordWrap(True)

        self.scan_table = QTableWidget(4, 6)
        self.scan_table.setHorizontalHeaderLabels(
            ["구분", "파일명", "시트", "자동 감지", "수정 시작행", "확인"]
        )
        self.scan_table.verticalHeader().setVisible(False)
        self.scan_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.scan_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.scan_table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.scan_table.horizontalHeader().setStretchLastSection(False)
        self.scan_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.scan_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.scan_table.setColumnWidth(0, 70)
        self.scan_table.setColumnWidth(2, 90)
        self.scan_table.setColumnWidth(3, 70)
        self.scan_table.setColumnWidth(4, 90)
        self.scan_table.setColumnWidth(5, 50)
        self.scan_table.setMouseTracking(True)

        def _scan_table_mouse_move(e):
            idx = self.scan_table.indexAt(e.pos())
            if idx.isValid() and idx.column() == 1:
                self.scan_table.setCursor(Qt.CursorShape.PointingHandCursor)
            else:
                self.scan_table.setCursor(Qt.CursorShape.ArrowCursor)
        self.scan_table.mouseMoveEvent = _scan_table_mouse_move

        header_h = self.scan_table.horizontalHeader().height()
        rows_h = sum(self.scan_table.rowHeight(i) for i in range(min(4, self.scan_table.rowCount())))
        frame_h = self.scan_table.frameWidth() * 2
        scroll_h = self.scan_table.horizontalScrollBar().sizeHint().height()
        extra = 10  # 여유 (필요하면 12~14로 올려)

        table_h = header_h + rows_h + frame_h + scroll_h + extra

        self.scan_table.setMinimumHeight(table_h)
        self.scan_table.setMaximumHeight(table_h)

        

        SCAN_TABLE_KINDS = ["신입생", "전입생", "전출생", "교직원"]

        for r, kind in enumerate(SCAN_TABLE_KINDS):
            # 0열: 구분 (편집 불가), 1~3열: 빈값
            for c, value in enumerate([kind, "", "", ""]):
                item = QTableWidgetItem(value)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.scan_table.setItem(r, c, item)

            # 수정 시작행 커스텀 +/- 위젯
            spin_widget = QWidget()
            spin_widget.setStyleSheet("background: transparent;")
            spin_layout = QHBoxLayout(spin_widget)
            spin_layout.setContentsMargins(0, 0, 0, 0)
            spin_layout.setSpacing(2)

            btn_minus = QPushButton("−")
            btn_minus.setFixedSize(20, 20)
            btn_minus.setStyleSheet("""
                QPushButton {
                    background: #F1F5F9;
                    border: 1px solid #D1D5DB;
                    border-radius: 5px;
                    font-size: 12px;
                    font-weight: 700;
                    color: #334155;
                    padding: 0px;
                    min-height: 0px;
                }
                QPushButton:hover {
                    background: #E2E8F0;
                }
            """)

            val_label = QLabel("-")
            val_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            val_label.setFixedWidth(20)
            val_label.setStyleSheet("font-weight: 700; color: #94A3B8; font-size: 12px;")

            btn_plus = QPushButton("+")
            btn_plus.setFixedSize(20, 20)
            btn_plus.setStyleSheet("""
                QPushButton {
                    background: #F1F5F9;
                    border: 1px solid #D1D5DB;
                    border-radius: 5px;
                    font-size: 12px;
                    font-weight: 700;
                    color: #334155;
                    padding: 0px;
                    min-height: 0px;
                }
                QPushButton:hover {
                    background: #E2E8F0;
                }
            """)

            def make_minus(lbl):
                def _():
                    try:
                        v = int(lbl.text())
                        if v > 1:
                            lbl.setText(str(v - 1))
                            lbl.setStyleSheet("font-weight: 700; color: #0F172A; font-size: 12px;")
                    except Exception:
                        pass
                return _

            def make_plus(lbl):
                def _():
                    try:
                        v = int(lbl.text())
                        lbl.setText(str(v + 1))
                    except Exception:
                        lbl.setText("1")
                    lbl.setStyleSheet("font-weight: 700; color: #0F172A; font-size: 12px;")
                return _

            btn_minus.clicked.connect(make_minus(val_label))
            btn_plus.clicked.connect(make_plus(val_label))

            spin_layout.addStretch()
            spin_layout.addWidget(btn_minus)
            spin_layout.addWidget(val_label)
            spin_layout.addWidget(btn_plus)
            spin_layout.addStretch()

            self.scan_table.setCellWidget(r, 4, spin_widget)

            # 확인 체크박스
            chk_widget = QWidget()
            chk_layout = QHBoxLayout(chk_widget)
            chk_layout.setContentsMargins(0, 0, 0, 0)
            chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chk = QCheckBox()
            chk_layout.addWidget(chk)
            self.scan_table.setCellWidget(r, 5, chk_widget)

        viewer_card = make_card()
        viewer_card.setSizePolicy(
            viewer_card.sizePolicy().horizontalPolicy(),
            __import__('PyQt6.QtWidgets', fromlist=['QSizePolicy']).QSizePolicy.Policy.Minimum
        )
        viewer_layout = QVBoxLayout(viewer_card)
        viewer_layout.setContentsMargins(14, 8, 14, 8)
        viewer_layout.setSpacing(0)

        # ── 헤더 행 (항상 표시) ──────────────────────────────
        header_row = QHBoxLayout()
        header_row.setContentsMargins(0, 0, 0, 0)
        header_row.setSpacing(8)

        viewer_title = QLabel("선택 파일 전체 검수")
        viewer_title.setObjectName("CardTitle")

        self.btn_toggle_viewer = QPushButton("펼치기 ▾")
        self.btn_toggle_viewer.setObjectName("GhostButton")
        self.btn_toggle_viewer.setCheckable(True)
        self.btn_toggle_viewer.setFixedHeight(28)
        self.btn_toggle_viewer.setStyleSheet("""
            QPushButton {
                background: transparent;
                border: 1px solid #D1D5DB;
                border-radius: 8px;
                padding: 2px 10px;
                font-size: 12px;
                font-weight: 700;
                color: #475569;
            }
            QPushButton:hover { background: #F8FAFC; }
        """)

        header_row.addWidget(viewer_title)
        header_row.addStretch()
        header_row.addWidget(self.btn_toggle_viewer)

        # ── 접힐 내용 영역 ────────────────────────────────────
        self.viewer_body = QWidget()
        body_layout = QVBoxLayout(self.viewer_body)
        body_layout.setContentsMargins(0, 6, 0, 0)
        body_layout.setSpacing(6)

        self.preview_file_info = QLabel("파일: - | 시트: - | 헤더행: - | 시작행: -")
        self.preview_file_info.setObjectName("Muted")

        filter_row = QHBoxLayout()
        filter_row.setSpacing(10)

        self.preview_search_input = QLineEdit()
        self.preview_search_input.setPlaceholderText("이름, 반, 학년 등 검색")

        self.btn_blank_only = QPushButton("빈 행/결측만")
        self.btn_blank_only.setCheckable(True)
        self.btn_blank_only.setObjectName("GhostButton")

        self.btn_issue_only = QPushButton("의심 행만")
        self.btn_issue_only.setCheckable(True)
        self.btn_issue_only.setObjectName("GhostButton")

        self.btn_dup_only = QPushButton("동명이인만")
        self.btn_dup_only.setCheckable(True)
        self.btn_dup_only.setObjectName("GhostButton")

        filter_row.addWidget(self.preview_search_input, 1)
        filter_row.addWidget(self.btn_blank_only)
        filter_row.addWidget(self.btn_issue_only)
        filter_row.addWidget(self.btn_dup_only)

        self.preview_warning_label = QLabel("검수할 파일을 위 표에서 선택하세요.")
        self.preview_warning_label.setObjectName("Muted")
        self.preview_warning_label.setWordWrap(True)

        self.preview_table = QTableWidget()
        self.preview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.preview_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.preview_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.preview_table.horizontalHeader().setStretchLastSection(False)
        self.preview_table.verticalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_table.verticalHeader().setMinimumWidth(36)
        self.preview_table.verticalHeader().setFixedWidth(40)
        self.preview_table.setMinimumHeight(600)

        body_layout.addWidget(self.preview_file_info)
        body_layout.addLayout(filter_row)
        body_layout.addWidget(self.preview_warning_label)
        body_layout.addWidget(self.preview_table)

        self.viewer_body.setVisible(False)  # 기본 접힘

        def _toggle_viewer():
            expanded = self.btn_toggle_viewer.isChecked()
            self.viewer_body.setVisible(expanded)
            self.btn_toggle_viewer.setText("접기 ▴" if expanded else "펼치기 ▾")
            viewer_card.setMaximumHeight(16777215)  # 제한 해제
            viewer_card.adjustSize()

        self.btn_toggle_viewer.clicked.connect(_toggle_viewer)

        viewer_layout.addLayout(header_row)
        viewer_layout.addWidget(self.viewer_body)

        # ── 학교 구분 자동 판별 실패 시에만 표시되는 예외 처리 UI ──────────
        self.school_kind_warn_label = QLabel("⚠  학교 구분을 자동으로 판별하지 못했습니다. 직접 선택한 뒤 실행해 주세요.")
        self.school_kind_warn_label.setStyleSheet(
            "color: #D97706; font-weight: 600; font-size: 12px;"
        )
        self.school_kind_warn_label.setWordWrap(True)
        self.school_kind_warn_label.hide()

        school_kind_row = QHBoxLayout()
        school_kind_row.setSpacing(8)
        school_kind_row.setContentsMargins(0, 0, 0, 0)

        school_kind_row_label = QLabel("학교 구분")
        school_kind_row_label.setObjectName("Muted")

        self.school_kind_combo = QComboBox()
        self.school_kind_combo.addItems(["초등부", "중등부", "고등부", "기타(빈칸)"])
        self.school_kind_combo.setFixedWidth(130)

        school_kind_row.addWidget(school_kind_row_label)
        school_kind_row.addWidget(self.school_kind_combo)
        school_kind_row.addStretch()

        self.school_kind_row_widget = QWidget()
        self.school_kind_row_widget.setLayout(school_kind_row)
        self.school_kind_row_widget.hide()

        # scan_result_widget 하위 호환용 더미 (참조 보호)
        self.scan_result_widget = QWidget()
        self.scan_result_widget.setMaximumHeight(0)
        self.scan_result_summary = QLabel("")
        self.scan_result_label = QLabel("")

        outer.addLayout(head_row)
        outer.addSpacing(2)
        outer.addWidget(self.scan_message)
        outer.addSpacing(2)
        outer.addWidget(self.scan_table)
        outer.addSpacing(2)
        outer.addWidget(self.school_kind_warn_label)
        outer.addWidget(self.school_kind_row_widget)
        outer.addSpacing(2)
        outer.addWidget(viewer_card)

        self.scan_preview_data = {}
        self.current_preview_kind = None
        self.preview_file_info.setText("파일: - | 시트: - | 헤더행: - | 시작행: -")
        self.preview_warning_label.setText("학교를 선택하고 스캔을 실행해 주세요.")
        self.preview_table.setRowCount(0)
        self.preview_table.setColumnCount(0)

        self.scan_table.itemSelectionChanged.connect(self.on_scan_table_row_selected)
        self.scan_table.cellDoubleClicked.connect(self.on_scan_table_double_clicked)
        self.preview_search_input.textChanged.connect(self.refresh_preview_table)
        self.btn_blank_only.toggled.connect(self.refresh_preview_table)
        self.btn_issue_only.toggled.connect(self.refresh_preview_table)
        self.btn_dup_only.toggled.connect(self.refresh_preview_table)

            
        QTimer.singleShot(0, lambda: self.fit_scan_table_height(8, 28))

        return card

    def _build_run_box(self):
        outer_card = make_card()
        outer_layout = QVBoxLayout(outer_card)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.setSpacing(0)

        content = QWidget()
        root = QVBoxLayout(content)
        root.setContentsMargins(18, 18, 18, 18)
        root.setSpacing(14)

        title = QLabel("실행 / 결과")
        title.setObjectName("CardTitle")

        desc = QLabel("스캔 검수가 끝난 뒤 작업을 실행하고 결과 파일을 확인합니다.")
        desc.setObjectName("Muted")

        memo_label = QLabel("메모")
        memo_label.setObjectName("Muted")

        self.result_note = QTextEdit()
        self.result_note.setPlaceholderText("이번 작업 관련 비고를 적어 주세요.")
        self.result_note.setFixedHeight(72)

        top_row = QHBoxLayout()
        top_row.setSpacing(8)

        self.btn_run = QPushButton("작업 실행")
        self.btn_run.setObjectName("PrimaryButton")
        self.btn_show_run_log = QPushButton("실행 로그 보기")
        self.btn_show_run_log.setObjectName("GhostButton")

        self.run_status_label = QLabel("실행 전")
        self.run_status_label.setStyleSheet(_STATUS_STYLE_IDLE)
        self.run_status_label.setFixedWidth(90)
        self.run_status_label.setFixedHeight(40)
        self.run_status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.btn_goto_notice_tab = QPushButton("→")
        self.btn_goto_notice_tab.setObjectName("PrimaryButton")
        self.btn_goto_notice_tab.setEnabled(False)
        self.btn_goto_notice_tab.setFixedSize(40, 40)
        self.btn_goto_notice_tab.setStyleSheet("""
            QPushButton {
                background: #2563EB; color: white;
                border-radius: 10px; font-size: 18px; font-weight: 700;
                padding: 0px;
            }
            QPushButton:hover { background: #1D4ED8; }
            QPushButton:disabled { background: #93C5FD; }
        """)
        self.btn_goto_notice_tab.setVisible(False)

        top_row.addWidget(self.btn_run)
        top_row.addWidget(self.btn_show_run_log)
        top_row.addWidget(self.run_status_label)
        top_row.addWidget(self.btn_goto_notice_tab)
        top_row.addStretch()

        self.run_info = QLabel("먼저 스캔을 통과해야 실행할 수 있습니다.")
        self.run_info.setObjectName("Muted")
        self.run_info.setWordWrap(True)

        self.run_hold_warning = QLabel("")
        self.run_hold_warning.setStyleSheet("""
            background: #FEF3C7;
            border: 1px solid #FDE68A;
            border-radius: 10px;
            padding: 10px 14px;
            color: #92400E;
            font-weight: 600;
        """)
        self.run_hold_warning.setWordWrap(True)
        self.run_hold_warning.hide()

        # ===== 파일 선택 바 =====
        # 하위 호환용 더미들
        self.run_file_title = QLabel("")
        self.run_file_title.setMaximumHeight(0)
        self.run_file_title.hide()
        self.run_file_combo = QComboBox()
        self.run_file_combo.setMaximumHeight(0)
        self.run_file_combo.hide()
        self.btn_open_file = QPushButton()
        self.btn_open_file.setMaximumHeight(0)
        self.btn_open_file.hide()
        self.output_dir = QLineEdit()
        self.output_dir.setMaximumHeight(0)
        self.output_dir.hide()

        file_bar_card = make_card()
        file_bar_layout = QVBoxLayout(file_bar_card)
        file_bar_layout.setContentsMargins(14, 10, 14, 10)
        file_bar_layout.setSpacing(6)

        file_bar_top = QHBoxLayout()
        file_bar_top.setContentsMargins(0, 0, 0, 0)
        file_bar_top.setSpacing(10)

        file_label = QLabel("출력 파일")
        file_label.setObjectName("CardTitle")

        self.btn_open_folder = QPushButton("폴더 열기")
        self.btn_open_folder.setObjectName("GhostButton")

        file_bar_top.addWidget(file_label)
        file_bar_top.addStretch()
        file_bar_top.addWidget(self.btn_open_folder)
        file_bar_layout.addLayout(file_bar_top)

        # 파일명 표 (더블클릭으로 열기, 스캔 탭 스타일)
        self.run_file_list_widget = QTableWidget(0, 1)
        self.run_file_list_widget.setHorizontalHeaderLabels(["파일명"])
        self.run_file_list_widget.horizontalHeader().setStretchLastSection(True)
        self.run_file_list_widget.verticalHeader().setVisible(False)
        self.run_file_list_widget.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.run_file_list_widget.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.run_file_list_widget.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.run_file_list_widget.setShowGrid(False)
        self.run_file_list_widget.setAlternatingRowColors(True)
        self.run_file_list_widget.setMouseTracking(True)
        self.run_file_list_widget.setFixedHeight(80)
        self.run_file_list_widget.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                font-size: 13px;
                color: #1D4ED8;
                background: white;
            }
            QTableWidget::item { padding: 6px 8px; }
            QTableWidget::item:hover { background: #EFF6FF; }
            QHeaderView::section {
                background: #F8FAFC;
                border: none;
                border-bottom: 1px solid #E5E7EB;
                padding: 6px 8px;
                font-size: 12px;
                font-weight: 600;
                color: #6B7280;
            }
        """)

        def _run_file_mouse_move(e):
            idx = self.run_file_list_widget.indexAt(e.pos())
            if idx.isValid():
                self.run_file_list_widget.setCursor(Qt.CursorShape.PointingHandCursor)
            else:
                self.run_file_list_widget.setCursor(Qt.CursorShape.ArrowCursor)
        self.run_file_list_widget.mouseMoveEvent = _run_file_mouse_move
        self.run_file_list_widget.cellDoubleClicked.connect(
            lambda row, col: self._open_run_file_signal.emit(
                self.run_file_list_widget.item(row, 0).text()
                if self.run_file_list_widget.item(row, 0) else ""
            )
        )
        file_bar_layout.addWidget(self.run_file_list_widget)

        # ===== 상태 요약 =====
        summary_card = make_card()
        summary_layout = QVBoxLayout(summary_card)
        summary_layout.setContentsMargins(14, 14, 14, 14)
        summary_layout.setSpacing(10)

        summary_title = QLabel("상태 요약")
        summary_title.setObjectName("CardTitle")

        self.run_summary_grid = QGridLayout()
        self.run_summary_grid.setHorizontalSpacing(12)
        self.run_summary_grid.setVerticalSpacing(8)

        self.sum_school = QLabel("-")
        self.sum_year = QLabel("-")
        self.sum_freshmen = QLabel("-")
        self.sum_teacher = QLabel("-")
        self.sum_transfer = QLabel("-")
        self.sum_withdraw = QLabel("-")
        self.sum_transfer_check = QLabel("-")
        self.sum_withdraw_check = QLabel("-")

        summary_items = [
            ("학교", self.sum_school),
            ("학년도", self.sum_year),
            ("신입생 입력", self.sum_freshmen),
            ("교사 입력", self.sum_teacher),
            ("전입 처리", self.sum_transfer),
            ("전출 처리", self.sum_withdraw),
            ("전입 합계", self.sum_transfer_check),
            ("전출 합계", self.sum_withdraw_check),
        ]

        for i, (label_text, value_widget) in enumerate(summary_items):
            lbl = QLabel(label_text)
            lbl.setObjectName("Muted")
            value_widget.setStyleSheet("font-weight: 800; color: #111827;")
            self.run_summary_grid.addWidget(lbl, i // 2, (i % 2) * 2)
            self.run_summary_grid.addWidget(value_widget, i // 2, (i % 2) * 2 + 1)

        summary_layout.addWidget(summary_title)
        summary_layout.addLayout(self.run_summary_grid)

        # ===== 시트 뷰어 =====
        viewer_card = make_card()
        viewer_card.setMinimumHeight(500)
        viewer_layout = QVBoxLayout(viewer_card)
        viewer_layout.setContentsMargins(14, 14, 14, 14)
        viewer_layout.setSpacing(10)

        viewer_title = QLabel("파일 뷰어")
        viewer_title.setObjectName("CardTitle")

        search_row = QHBoxLayout()
        search_row.setSpacing(8)

        self.run_view_search = QLineEdit()
        self.run_view_search.setPlaceholderText("현재 시트에서 검색")

        self.btn_run_dup_only = QPushButton("동명이인만")
        self.btn_run_dup_only.setCheckable(True)
        self.btn_run_dup_only.setObjectName("GhostButton")

        self.run_preview_info = QLabel("시트: - | 행 수: -")
        self.run_preview_info.setObjectName("Muted")

        search_row.addWidget(self.run_view_search, 1)
        search_row.addWidget(self.btn_run_dup_only)

        self.run_sheet_tabs = QTabWidget()

        viewer_layout.addWidget(viewer_title)
        viewer_layout.addLayout(search_row)
        viewer_layout.addWidget(self.run_preview_info)
        viewer_layout.addWidget(self.run_sheet_tabs, 1)

        # (export_row 제거됨)

        # run_result_widget 하위 호환용 더미
        self.run_result_widget = QWidget()
        self.run_result_widget.setMaximumHeight(0)
        self.run_result_label = QLabel("")
        self.run_result_summary = QLabel("")

        # btn_export_log 하위 호환용 더미
        self.btn_export_log = QPushButton()
        self.btn_export_log.setMaximumHeight(0)
        self.btn_export_log.hide()
        self.export_path = QLineEdit()
        self.export_path.setMaximumHeight(0)
        self.export_path.hide()

        root.addWidget(title)
        root.addWidget(desc)
        root.addWidget(memo_label)
        root.addWidget(self.result_note)
        root.addLayout(top_row)
        root.addWidget(self.run_info)
        root.addWidget(self.run_hold_warning)
        root.addWidget(summary_card)
        root.addWidget(file_bar_card)
        root.addWidget(viewer_card, 1)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setWidget(content)

        outer_layout.addWidget(scroll)

        self.run_file_combo.clear()
        self.run_file_combo.currentTextChanged.connect(self.load_selected_run_file)
        self.run_view_search.textChanged.connect(self.filter_current_run_sheet)
        self.run_sheet_tabs.currentChanged.connect(self.update_run_preview_info)
        self.btn_run_dup_only.toggled.connect(self.filter_current_run_sheet)

        self.run_file_title.setText("실행 전")
        self.run_preview_info.setText("시트: - | 행 수: -")
        self.output_dir.setText("")

        return outer_card

    def _build_notice_box(self):
        card = make_card()
        root = QVBoxLayout(card)
        root.setContentsMargins(18, 18, 18, 18)
        root.setSpacing(14)

        title = QLabel("안내문")
        title.setObjectName("CardTitle")

        desc = QLabel("학교명·학년도·개학일·도메인이 자동으로 반영되는 안내문 초안입니다.")
        desc.setObjectName("Muted")

        body = QHBoxLayout()
        body.setSpacing(12)

        self.notice_list = QListWidget()

        right = QVBoxLayout()
        right.setSpacing(10)

        self.notice_text = QPlainTextEdit()
        self.notice_text.setPlainText(
            "실행 완료 후 안내문이 자동으로 채워집니다.\n\n"
            "학교명·학년도·개학일·도메인이 자동으로 반영됩니다."
        )
        self.notice_text.setReadOnly(False)

        btn_row = QHBoxLayout()
        btn_row.addStretch()

        self.btn_copy_notice = QPushButton("복사")
        self.btn_copy_notice.setObjectName("GhostButton")
        self.btn_reset_notice = QPushButton("초기화")
        self.btn_reset_notice.setObjectName("GhostButton")

        btn_row.addWidget(self.btn_copy_notice)
        btn_row.addWidget(self.btn_reset_notice)

        # 하위 호환용 더미 (로직 참조 보호)
        self.btn_email_sent = QPushButton()
        self.btn_email_sent.setCheckable(True)
        self.btn_email_sent.setMaximumHeight(0)
        self.btn_email_sent.hide()
        self.btn_email_hold = QPushButton()
        self.btn_email_hold.setCheckable(True)
        self.btn_email_hold.setMaximumHeight(0)
        self.btn_email_hold.hide()
        self.email_log_status = QLabel("")
        self.email_log_status.setMaximumHeight(0)
        self.email_log_status.hide()
        self.btn_record_roster = QPushButton()
        self.btn_record_roster.setEnabled(False)
        self.btn_record_roster.setMaximumHeight(0)
        self.btn_record_roster.hide()

        right.addWidget(self.notice_text, 1)
        right.addLayout(btn_row)

        body.addWidget(self.notice_list, 1)
        body.addLayout(right, 3)

        root.addWidget(title)
        root.addWidget(desc)
        root.addLayout(body, 1)

        return card

    def _build_diff_run_box(self):
        card = make_card()
        root = QVBoxLayout(card)
        root.setContentsMargins(18, 18, 18, 18)
        root.setSpacing(14)

        title = QLabel("재학생 명단 비교")
        title.setObjectName("CardTitle")

        self.diff_caption = QLabel(
            "기준 명부와 학교 재학생 명단을 비교해 차이 항목을 정리합니다.\n"
            "학교 명단에만 있는 학생, 명부에만 있는 학생, 자동 판정이 어려운 학생을 구분해 결과 파일을 생성합니다."
        )
        self.diff_caption.setWordWrap(True)
        self.diff_caption.setObjectName("Muted")

        # target_year 입력
        year_row = QHBoxLayout()
        year_row.setSpacing(8)
        year_label = QLabel("비교 학년도")
        year_label.setObjectName("Muted")
        year_label.setFixedWidth(72)
        self.diff_target_year = QSpinBox()
        self.diff_target_year.setRange(2020, 2040)
        self.diff_target_year.setValue(QDate.currentDate().year())
        self.diff_target_year.setFixedWidth(100)
        self.diff_target_year.setFixedHeight(36)
        year_row.addWidget(year_label)
        year_row.addWidget(self.diff_target_year)
        year_row.addStretch()

        btn_row = QHBoxLayout()
        self.btn_run_diff = QPushButton("명부 비교 실행")
        self.btn_run_diff.setObjectName("PrimaryButton")
        self.btn_show_diff_log = QPushButton("실행 로그 보기")
        self.btn_show_diff_log.setObjectName("GhostButton")
        self.diff_status_label = QLabel("실행 전")
        self.diff_status_label.setStyleSheet(_STATUS_STYLE_IDLE)
        btn_row.addWidget(self.btn_run_diff)
        btn_row.addWidget(self.btn_show_diff_log)
        btn_row.addWidget(self.diff_status_label)
        btn_row.addStretch()

        root.addWidget(title)
        root.addWidget(self.diff_caption)
        root.addLayout(year_row)
        root.addLayout(btn_row)

        return card

    def _build_diff_result_box(self):
        card = make_card()
        root = QVBoxLayout(card)
        root.setContentsMargins(18, 18, 18, 18)
        root.setSpacing(14)

        title = QLabel("비교 결과")
        title.setObjectName("CardTitle")

        # ── 사전 점검 결과 ─────────────────────────────────────
        self.diff_scan_info = QLabel("")
        self.diff_scan_info.setObjectName("Muted")
        self.diff_scan_info.setWordWrap(True)
        self.diff_scan_info.hide()

        # ── 요약 수치 카드 4개 ─────────────────────────────────
        stat_row = QHBoxLayout()
        stat_row.setSpacing(10)

        def _stat_card(title_text, desc_text):
            f = make_card()
            fl = QVBoxLayout(f)
            fl.setContentsMargins(12, 10, 12, 10)
            fl.setSpacing(2)
            t = QLabel(title_text)
            t.setObjectName("Muted")
            v = QLabel("-")
            v.setStyleSheet("font-size: 18px; font-weight: 800; color: #111827;")
            d = QLabel(desc_text)
            d.setObjectName("Muted")
            d.setWordWrap(True)
            fl.addWidget(t)
            fl.addWidget(v)
            fl.addWidget(d)
            return f, v

        f1, self.diff_roster_only   = _stat_card("명부 누락",  "명부에만 있음")
        f2, self.diff_matched       = _stat_card("정상 유지",  "양쪽 일치")
        f3, self.diff_compare_only  = _stat_card("신규 포함",  "현재 명단에만 있음")
        f4, self.diff_unresolved    = _stat_card("확인 필요",  "자동 판정 불가")

        for f in [f1, f2, f3, f4]:
            stat_row.addWidget(f, 1)

        # ── 3열 집합 뷰어 ──────────────────────────────────────
        viewer_row = QHBoxLayout()
        viewer_row.setSpacing(10)

        def _col_widget(col_title, col_desc):
            w = QWidget()
            vl = QVBoxLayout(w)
            vl.setContentsMargins(0, 0, 0, 0)
            vl.setSpacing(4)
            ht = QLabel(col_title)
            ht.setStyleSheet("font-weight: 700; font-size: 13px; color: #0F172A;")
            hd = QLabel(col_desc)
            hd.setObjectName("Muted")
            hd.setWordWrap(True)
            tbl = QTableWidget()
            tbl.setColumnCount(3)
            tbl.setHorizontalHeaderLabels(["학년", "반", "이름"])
            tbl.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
            tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
            tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
            tbl.setColumnWidth(0, 44)
            tbl.setColumnWidth(1, 44)
            tbl.verticalHeader().setVisible(False)
            tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            vl.addWidget(ht)
            vl.addWidget(hd)
            vl.addWidget(tbl, 1)
            return w, tbl

        col1, self.diff_tbl_roster_only  = _col_widget("명부 누락",
            "명부에는 있으나 현재 명단에 없음")
        col2, self.diff_tbl_matched      = _col_widget("정상 유지",
            "양쪽 명단 모두 존재")
        col3, self.diff_tbl_compare_only = _col_widget("신규 포함",
            "현재 명단에는 있으나 명부에 없음")

        viewer_row.addWidget(col1, 1)
        viewer_row.addWidget(col2, 1)
        viewer_row.addWidget(col3, 1)

        # ── 확인 필요 박스 ─────────────────────────────────────
        unresolved_label = QLabel("확인 필요")
        unresolved_label.setStyleSheet("font-weight: 700; font-size: 13px; color: #0F172A;")
        unresolved_desc = QLabel("자동 판정 불가 — 직접 확인이 필요한 항목")
        unresolved_desc.setObjectName("Muted")

        self.diff_tbl_unresolved = QTableWidget()
        self.diff_tbl_unresolved.setColumnCount(4)
        self.diff_tbl_unresolved.setHorizontalHeaderLabels(["학년", "반", "이름", "사유"])
        self.diff_tbl_unresolved.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.diff_tbl_unresolved.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.diff_tbl_unresolved.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.diff_tbl_unresolved.setColumnWidth(0, 44)
        self.diff_tbl_unresolved.setColumnWidth(1, 44)
        self.diff_tbl_unresolved.verticalHeader().setVisible(False)
        self.diff_tbl_unresolved.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.diff_tbl_unresolved.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.diff_tbl_unresolved.setMaximumHeight(140)

        # ── 자동 분류 요약 + 생성 파일 ────────────────────────
        self.diff_summary = QLabel("명단 비교 실행 버튼을 눌러 주세요.")
        self.diff_summary.setWordWrap(True)
        self.diff_summary.setObjectName("Muted")

        files_row = QHBoxLayout()
        files_row.setSpacing(8)
        files_label = QLabel("생성 파일")
        files_label.setObjectName("Muted")
        files_label.setFixedWidth(52)
        self.diff_outputs = QListWidget()
        self.diff_outputs.setMaximumHeight(60)
        self.diff_outputs.addItems(["전입생 명단.xlsx", "전출생 명단.xlsx"])
        self.diff_output_dir = QLineEdit()
        self.diff_output_dir.setPlaceholderText("저장 위치")
        self.diff_output_dir.setReadOnly(True)
        files_row.addWidget(files_label)
        files_row.addWidget(self.diff_outputs, 1)
        files_row.addWidget(self.diff_output_dir, 1)

        root.addWidget(title)
        root.addWidget(self.diff_scan_info)
        root.addLayout(stat_row)
        root.addLayout(viewer_row, 1)
        root.addWidget(unresolved_label)
        root.addWidget(unresolved_desc)
        root.addWidget(self.diff_tbl_unresolved)
        root.addWidget(self.diff_summary)
        root.addLayout(files_row)

        return card

    def _init_scan_preview_sample_data(self):
        self.scan_preview_data = {}

    def refresh_preview_table(self):
        kind = self.current_preview_kind
        if not kind:
            return

        data = self.scan_preview_data.get(kind)
        if not data:
            return

        headers = data["headers"]
        rows = data["rows"]
        issue_rows = data["issue_rows"]

        keyword = self.preview_search_input.text().strip().lower()
        blank_only = self.btn_blank_only.isChecked()
        issue_only = self.btn_issue_only.isChecked()
        dup_only   = self.btn_dup_only.isChecked()

        # 이름 컬럼 인덱스 탐색
        NAME_KEYWORDS = ["이름", "성명", "학생이름", "선생님이름", "교사이름"]
        name_col = None
        for kw in NAME_KEYWORDS:
            for i, h in enumerate(headers):
                if kw in str(h):
                    name_col = i
                    break
            if name_col is not None:
                break

        # 학년 컬럼 인덱스 탐색
        grade_col = None
        for i, h in enumerate(headers):
            if "학년" in str(h):
                grade_col = i
                break

        # 동명이인 행 인덱스 사전 계산 — 학년+이름 기준 (학년 컬럼 없으면 이름만)
        # suffix(A/B/AA...) 제거한 원본 이름 기준으로 판정
        dup_indices: set = set()
        if dup_only and name_col is not None:
            import re as _re
            from collections import Counter
            def _base_name(nm: str) -> str:
                if _re.search(r"[\uAC00-\uD7A3]", nm):
                    return _re.sub(r"[A-Z]+$", "", nm).strip()
                return nm
            key_count: Counter = Counter()
            for row in rows:
                if name_col < len(row):
                    nm = _base_name(str(row[name_col]).strip())
                    if not nm:
                        continue
                    grade = str(row[grade_col]).strip() if (grade_col is not None and grade_col < len(row)) else ""
                    key_count[(grade, nm)] += 1
            for idx, row in enumerate(rows):
                if name_col < len(row):
                    nm = _base_name(str(row[name_col]).strip())
                    if not nm:
                        continue
                    grade = str(row[grade_col]).strip() if (grade_col is not None and grade_col < len(row)) else ""
                    if key_count[(grade, nm)] >= 2:
                        dup_indices.add(idx)

        filtered = []
        for idx, row in enumerate(rows):
            row_values = [str(v) for v in row]

            if keyword:
                joined = " ".join(row_values).lower()
                if keyword not in joined:
                    continue

            # 핵심 컬럼(이름, 학년)이 비어있는 경우만 blank로 판정
            key_cols = [c for c in [name_col, grade_col] if c is not None and c < len(row)]
            if key_cols:
                has_blank = any(str(row[c]).strip() == "" for c in key_cols)
                all_blank = all(str(row[c]).strip() == "" for c in key_cols)
            else:
                has_blank = any(str(v).strip() == "" for v in row)
                all_blank = all(str(v).strip() == "" for v in row)

            if blank_only and not (has_blank or all_blank):
                continue

            if issue_only and idx not in issue_rows:
                continue

            if dup_only and idx not in dup_indices:
                continue

            filtered.append((idx, row))

        self.preview_table.clear()
        self.preview_table.setColumnCount(len(headers))
        self.preview_table.setHorizontalHeaderLabels(headers)
        self.preview_table.setRowCount(len(filtered))
        self.preview_table.verticalHeader().setDefaultSectionSize(28)

        data = self.scan_preview_data.get(kind, {})
        data_start_row = data.get("data_start_row", 1) or 1

        for r, (src_idx, row) in enumerate(filtered):
            excel_row = data_start_row + src_idx
            header_item = QTableWidgetItem(str(excel_row))
            header_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.preview_table.setVerticalHeaderItem(r, header_item)

        for r, (src_idx, row) in enumerate(filtered):
            key_cols = [c for c in [name_col, grade_col] if c is not None and c < len(row)]
            if key_cols:
                has_blank = any(str(row[c]).strip() == "" for c in key_cols)
                all_blank = all(str(row[c]).strip() == "" for c in key_cols)
            else:
                # 이름/학년 컬럼을 특정할 수 없으면 강조 없음
                has_blank = False
                all_blank = False
            is_issue  = src_idx in issue_rows
            is_dup    = src_idx in dup_indices

            for c, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                self.preview_table.setItem(r, c, item)

                if all_blank:
                    item.setBackground(QColor("#FDEAEA"))
                elif has_blank:
                    item.setBackground(QColor("#FFF6D6"))
                elif is_issue:
                    item.setBackground(QColor("#FFE7CC"))
                elif is_dup:
                    # #15 동명이인 노란색 강조
                    item.setBackground(QColor("#FEF9C3"))

        self.preview_table.resizeColumnsToContents()

    def on_scan_table_double_clicked(self, row: int, col: int):
        """스캔 표 행 더블클릭 → 해당 파일 열기 시그널."""
        item = self.scan_table.item(row, 1)  # 1열: 파일명
        if item and item.text().strip():
            self._open_scan_file_signal.emit(item.text().strip())

    def on_scan_table_row_selected(self):
        """스캔 표 행 클릭 → 미리보기 갱신."""
        row = self.scan_table.currentRow()
        if row < 0:
            return
        kind_item = self.scan_table.item(row, 0)
        if not kind_item:
            return
        kind = kind_item.text().strip()
        # Bug 3 수정: scan_preview_data에 해당 kind가 없으면 무시
        if kind not in self.scan_preview_data:
            return
        self.load_selected_preview(kind)

    def load_selected_preview(self, kind: str = ""):
        """kind를 받아 scan_preview_data에서 데이터를 꺼내 미리보기를 갱신."""
        if not kind:
            return
        data = self.scan_preview_data.get(kind)
        if not data:
            return

        # Bug 4 수정: scan_main이 rows=[]로 반환하므로, 비어 있으면 실제 파일에서 로드
        if not data.get("rows"):
            try:
                from core.scan_main import load_preview_rows
                from pathlib import Path as _Path
                _KIND_KEY = {
                    "신입생": "freshmen",
                    "전입생": "transfer",
                    "전출생": "withdraw",
                    "교직원": "teacher",
                }
                file_name = data.get("file_name", "")
                input_dir = self._get_scan_input_dir()
                if file_name and input_dir is not None:
                    file_path = _Path(str(input_dir)) / file_name
                    if file_path.exists():
                        kind_key = _KIND_KEY.get(kind, kind)
                        rows = load_preview_rows(
                            xlsx_path=file_path,
                            kind=kind_key,
                            header_row=data.get("header_row", 1),
                            data_start_row=data.get("data_start_row", 2),
                            limit=3000,
                        )
                        data["rows"] = rows
                        self.scan_preview_data[kind] = data
            except Exception:
                pass

        self.current_preview_kind = kind
        self.preview_file_info.setText(
            f"파일: {data['file_name']} | 시트: {data['sheet_name']} | "
            f"헤더행: {data['header_row']} | 시작행: {data['data_start_row']}"
        )
        raw_warning = data["warning"]
        # 개발자 로그 prefix 제거 후 표시
        clean_lines = []
        for line in raw_warning.splitlines():
            line = re.sub(r"^\[(ERROR|WARN|INFO|OK|DONE)\]\s*", "", line).strip()
            if line:
                clean_lines.append(line)
        self.preview_warning_label.setText("\n".join(clean_lines) if clean_lines else "이상 없음")
        self.refresh_preview_table()

    def _get_scan_input_dir(self):
        """MainWindow의 _last_scan_result에서 input_dir을 가져온다.
        MainTab -> QStackedWidget -> MainPage -> ... -> MainWindow 계층을 모두 탐색.
        """
        w = self.parent()
        visited = set()
        while w is not None:
            wid = id(w)
            if wid in visited:
                break
            visited.add(wid)
            result = getattr(w, "_last_scan_result", None)
            if result is not None:
                return getattr(result, "input_dir", None)
            w = w.parent()
        return None

    def _init_run_preview_sample_data(self):
        self.run_preview_data = {}


    def load_selected_run_file(self, file_name=None):
        file_name = (file_name or "").strip()
        if not file_name:
            return
        
        data = self.run_preview_data.get(file_name)
        if not data:
            return

        self.current_run_file = file_name
        summary = data.get("summary", {})
        input_counts = summary.get("input_counts", {})
        result_counts = summary.get("result_counts", {})
        checks = summary.get("checks", {})

        self.run_file_title.setText(file_name)
        self.sum_school.setText(summary.get("school_name", "-"))
        self.sum_year.setText(summary.get("year_str", "-"))
        self.sum_freshmen.setText(str(input_counts.get("freshmen", 0)))
        self.sum_teacher.setText(str(input_counts.get("teacher", 0)))

        self.sum_transfer.setText(
            f"{result_counts.get('transfer_done', 0)} 완료 / {result_counts.get('transfer_hold', 0)} 보류"
        )
        self.sum_withdraw.setText(
            f"{result_counts.get('withdraw_done', 0)} 완료 / {result_counts.get('withdraw_hold', 0)} 보류"
        )

        transfer_ok = checks.get("transfer_total_match", None)
        withdraw_ok = checks.get("withdraw_total_match", None)

        def check_text(val):
            if val is None:
                return "-"
            return "정상" if val else "불일치"

        def check_style(val):
            if val is None:
                return "font-weight: 800; color: #64748B;"
            return "font-weight: 800; color: #15803D;" if val else "font-weight: 800; color: #DC2626;"

        self.sum_transfer_check.setText(check_text(transfer_ok))
        self.sum_withdraw_check.setText(check_text(withdraw_ok))
        self.sum_transfer_check.setStyleSheet(check_style(transfer_ok))
        self.sum_withdraw_check.setStyleSheet(check_style(withdraw_ok))

        self.populate_run_sheet_tabs()

class MainPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        page = QWidget()
        root = QVBoxLayout(page)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(14)

        self.header = AppHeader()
        root.addWidget(self.header)

        # 상단 스텝바
        self.step_bar = StepBar()
        root.addWidget(self.step_bar)

        body = QHBoxLayout()
        body.setContentsMargins(0, 0, 0, 0)
        body.setSpacing(14)

        self.status_panel = StatusPanel()
        self.main_tab = MainTab()

        self.diff_widget = QWidget()
        diff_layout = QVBoxLayout(self.diff_widget)
        diff_layout.setContentsMargins(0, 0, 0, 0)
        diff_content = QWidget()
        diff_root = QVBoxLayout(diff_content)
        diff_root.setContentsMargins(14, 14, 14, 14)
        diff_root.setSpacing(14)
        diff_root.addWidget(self.main_tab._build_diff_run_box())
        diff_root.addWidget(self.main_tab._build_diff_result_box())
        diff_root.addStretch()
        diff_scroll = QScrollArea()
        diff_scroll.setWidgetResizable(True)
        diff_scroll.setFrameShape(QFrame.Shape.NoFrame)
        diff_scroll.setWidget(diff_content)
        diff_layout.addWidget(diff_scroll)

        self.content_stack = QStackedWidget()
        self.content_stack.addWidget(self.main_tab)    # index 0
        self.content_stack.addWidget(self.diff_widget) # index 1

        body.addWidget(self.status_panel)
        body.addWidget(self.content_stack, 1)

        body_wrap = QWidget()
        body_wrap.setLayout(body)
        root.addWidget(body_wrap, 1)

        scroll.setWidget(page)
        outer.addWidget(scroll)

    def set_mode(self, mode: str):
        if mode == "main":
            self.content_stack.setCurrentIndex(0)
            self.status_panel.progress_card_widget.show()
            self.header.btn_mode_main.setChecked(True)
            self.header.btn_mode_diff.setChecked(False)
        else:
            self.content_stack.setCurrentIndex(1)
            self.status_panel.progress_card_widget.hide()
            self.header.btn_mode_main.setChecked(False)
            self.header.btn_mode_diff.setChecked(True)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("리딩게이트 반이동 자동화")
        self.resize(1460, 940)

        self.stack = QStackedWidget()
        self.setup_page = SetupPage(self.goto_main_page)
        self.main_page = MainPage()

        self.worker_name = ""
        self.work_root = None
        self.work_date = None
        self.school_start_date = None
        self.selected_school = None
        self.roster_log_path = None       # 명단 파일 경로
        self._roster_col_map = {}         # 명단 열 매핑 설정
        self._pending_roster_log = False  # 실행 완료 후 명단 미기록 상태
        self._pending_history_entry = None

        self._school_names = []
        self._school_name_set = set()
        self._school_search_timer = QTimer(self)
        self._school_search_timer.setSingleShot(True)
        self._school_search_timer.setInterval(250)
        self._school_search_timer.timeout.connect(self._apply_school_search)
        self._pending_school_keyword = ""

        # 로그 팝업용 결과 보관
        self._last_scan_logs: list = []
        self._last_run_logs:  list = []
        self._last_diff_logs: list = []

        # 스캔/실행 결과 보관
        self._last_scan_result = None
        self._last_run_result = None  # 명단 기록 시 kind_flags 참조용

        # 파일 열기용 경로 보관
        self._current_output_files: list = []

        self.stack.addWidget(self.setup_page)  # index 0
        self.stack.addWidget(self.main_page)   # index 1
        self.setCentralWidget(self.stack)

        self._wire_signals()

        self.app_config = load_app_config()
        self._apply_saved_defaults()

        self.main_page.main_tab.notice_list.currentRowChanged.connect(self._on_notice_row_changed)
        self.main_page.main_tab.btn_copy_notice.clicked.connect(self._on_copy_notice)
        self.main_page.main_tab.btn_reset_notice.clicked.connect(self._on_reset_notice)

    def _wire_signals(self):
        self.main_page.status_panel.btn_select_school.clicked.connect(self.apply_selected_school)
        self.main_page.main_tab.tabs.currentChanged.connect(self.on_tab_changed)
        self.main_page.main_tab.btn_scan.clicked.connect(self.run_scan)
        self.main_page.main_tab.btn_run.clicked.connect(self.run_main)
        self.main_page.main_tab.btn_run_diff.clicked.connect(self.run_diff)

        # StepBar 콜백 연결
        sb = self.main_page.step_bar
        sb.set_callback(0, self.go_back_to_setup)
        sb.set_callback(1, self._reset_to_school_select)
        sb.set_callback(2, lambda: self._go_to_step(0))
        sb.set_callback(3, lambda: self._go_to_step(1))
        sb.set_callback(4, lambda: self._go_to_step(2))

        # 새 학교 시작 버튼
        self.main_page.status_panel.btn_new_school.clicked.connect(
            self._on_new_school_clicked
        )

        # 사이드바 명단 기록 버튼
        self.main_page.status_panel.btn_record_roster_sidebar.clicked.connect(
            self._on_record_roster
        )
        self.main_page.main_tab.btn_scan.clicked.connect(self.run_scan)
        self.main_page.main_tab.btn_run.clicked.connect(self.run_main)
        self.main_page.main_tab.btn_run_diff.clicked.connect(self.run_diff)

        # 로그 팝업
        self.main_page.main_tab.btn_show_scan_log.clicked.connect(
            lambda: self._show_log_dialog("스캔 로그", self._last_scan_logs)
        )
        self.main_page.main_tab.btn_show_run_log.clicked.connect(
            lambda: self._show_log_dialog("실행 로그", self._last_run_logs)
        )
        self.main_page.main_tab.btn_show_diff_log.clicked.connect(
            lambda: self._show_log_dialog("명단 비교 로그", self._last_diff_logs)
        )

        # 스캔 표 더블클릭 → 파일 열기
        self.main_page.main_tab._open_scan_file_signal.connect(self._open_scan_file_by_name)
        self.main_page.main_tab._open_run_file_signal.connect(self._on_run_file_double_clicked)

        # 스캔 완료 후 실행 단계 이동 — 체크박스 확인 후
        self.main_page.main_tab.btn_goto_run_tab.clicked.connect(
            self._on_goto_run_tab
        )

        # 실행 완료 후 안내문 단계 이동
        self.main_page.main_tab.btn_goto_notice_tab.clicked.connect(
            lambda: self._go_to_step(2)
        )

        # 파일/폴더 열기
        self.main_page.main_tab.btn_open_file.clicked.connect(self._open_selected_file)
        self.main_page.main_tab.btn_open_folder.clicked.connect(self._open_output_folder)

        # 찾아보기
        self.setup_page.btn_browse_path.clicked.connect(self._browse_work_root)
        self.setup_page.btn_browse_roster_log.clicked.connect(self._browse_roster_log)
        self.setup_page.btn_save_defaults.clicked.connect(self._save_default_settings)
        self.setup_page.btn_load_defaults.clicked.connect(self._load_default_settings)

        # 학년도 아이디 규칙 수동 적용
        self.main_page.status_panel.btn_grade_map_apply.clicked.connect(
            self._on_grade_map_apply
        )

        # 안내문 탭 — 이메일 발송/보류/명단 기록
        self.main_page.main_tab.btn_email_sent.clicked.connect(self._on_email_sent_clicked)
        self.main_page.main_tab.btn_email_hold.clicked.connect(self._on_email_hold_clicked)
        self.main_page.main_tab.btn_record_roster.clicked.connect(self._on_record_roster)

        # 모드 전환
        self.main_page.header.btn_mode_main.clicked.connect(
            lambda: (self.main_page.set_mode("main"), self.sync_task_name())
        )
        self.main_page.header.btn_mode_diff.clicked.connect(
            lambda: (self.main_page.set_mode("diff"), self.sync_task_name())
        )

        # 학교 검색 입력/리스트 연결
        school_input = self.main_page.status_panel.school_input
        school_list = self.main_page.status_panel.school_result_list

        school_input.textEdited.connect(self.on_school_text_edited)
        school_input.textChanged.connect(self.validate_school_input)
        school_input.set_key_callback(self.handle_school_input_keypress)

        school_list.itemClicked.connect(self.on_school_item_clicked)
        school_list.itemDoubleClicked.connect(self.on_school_item_double_clicked)

        # 초기 버튼 상태
        self.main_page.status_panel.btn_select_school.setEnabled(False)
        self.main_page.main_tab.btn_scan.setEnabled(False)
        self.main_page.main_tab.btn_run.setEnabled(False)
        self.main_page.main_tab.btn_run_diff.setEnabled(False)

        self.sync_task_name()


    def _apply_saved_defaults(self):
        cfg = self.app_config or {}

        self.setup_page.work_root_input.setText(cfg.get("work_root", ""))
        self.setup_page.roster_log_input.setText(cfg.get("roster_log_path", ""))
        self.setup_page.worker_input.setText(cfg.get("worker_name", ""))

        school_start = cfg.get("school_start_date", "")
        if school_start:
            qd = QDate.fromString(school_start, "yyyy-MM-dd")
            if qd.isValid():
                self.setup_page.open_date_edit.setDate(qd)

        work_date = cfg.get("work_date", "")
        if work_date:
            qd = QDate.fromString(work_date, "yyyy-MM-dd")
            if qd.isValid():
                self.setup_page.work_date_edit.setDate(qd)

        # 마지막 세션 도착일 불러오기
        last_arrived = cfg.get("last_arrived_date", "")
        if last_arrived:
            qd = QDate.fromString(last_arrived, "yyyy-MM-dd")
            if qd.isValid():
                self.main_page.status_panel.email_arrived_date.setDate(qd)

        self.setup_page._refresh_start_button()

    def _save_default_settings(self):
        config = {
            "work_root": self.setup_page.work_root_input.text().strip(),
            "roster_log_path": self.setup_page.roster_log_input.text().strip(),
            "worker_name": self.setup_page.worker_input.text().strip(),
            "school_start_date": self.setup_page.open_date_edit.date().toString("yyyy-MM-dd"),
            "work_date": self.setup_page.work_date_edit.date().toString("yyyy-MM-dd"),
        }

        try:
            save_app_config(config)
            self.app_config = config
            QMessageBox.information(self, "저장 완료", "기본 설정을 저장했습니다.")
        except Exception as e:
            QMessageBox.warning(self, "저장 실패", f"기본 설정 저장 중 오류가 발생했습니다.\n{e}")

    def _load_default_settings(self):
        self.app_config = load_app_config()

        if not any(str(v).strip() for v in self.app_config.values()):
            QMessageBox.information(self, "불러오기", "저장된 기본 설정이 없습니다.")
            return

        self._apply_saved_defaults()
        QMessageBox.information(self, "불러오기 완료", "저장된 기본 설정을 불러왔습니다.")

    def on_tab_changed(self, idx: int):
        self.sync_task_name()

    def _on_goto_run_tab(self):
        from PyQt6.QtWidgets import QMessageBox
        tab = self.main_page.main_tab
        table = tab.scan_table
        unchecked_kinds = []
        for kind, row in self._SCAN_TABLE_KIND_ROW.items():
            spin_widget = table.cellWidget(row, 4)
            if spin_widget is None or not spin_widget.isEnabled():
                continue
            chk_widget = table.cellWidget(row, 5)
            if chk_widget is not None:
                chk = chk_widget.findChild(QCheckBox)
                if chk is not None and not chk.isChecked():
                    unchecked_kinds.append(kind)
        if unchecked_kinds:
            QMessageBox.warning(
                self, "시작 행 확인 필요",
                "다음 항목의 시작 행을 확인해 주세요:\n"
                + ", ".join(unchecked_kinds)
            )
            return
        self._go_to_step(1)

    def _go_to_step(self, idx: int):
        """결과 영역 버튼으로 특정 단계(인덱스)로 이동."""
        self.main_page.main_tab.show_step(idx)
        self.sync_task_name()

    def _reveal_scan_result(self, summary_text: str):
        tab = self.main_page.main_tab
        tab.scan_result_summary.setText(summary_text)
        tab.btn_goto_run_tab.setEnabled(True)
        tab.btn_goto_run_tab.setVisible(True)

    def _reveal_run_result(self, summary_text: str):
        tab = self.main_page.main_tab
        tab.run_result_summary.setText(summary_text)
        tab.btn_goto_notice_tab.setEnabled(True)
        tab.btn_goto_notice_tab.setVisible(True)


    def goto_main_page(self):
        from datetime import date as _date
        from PyQt6.QtWidgets import QMessageBox
        from engine import inspect_work_root

        work_root_text = self.setup_page.work_root_input.text().strip()
        worker_text = self.setup_page.worker_input.text().strip()

        if not work_root_text:
            QMessageBox.warning(self, "입력 확인", "작업 폴더를 입력하세요.")
            return

        work_root_path = Path(work_root_text)
        if not work_root_path.exists() or not work_root_path.is_dir():
            QMessageBox.warning(self, "입력 확인", f"작업 폴더를 찾을 수 없습니다.\n{work_root_text}")
            return

        if not worker_text:
            QMessageBox.warning(self, "입력 확인", "작업자 이름을 입력하세요.")
            return

        inspect = inspect_work_root(work_root_path)
        if not inspect.get("ok", False):
            errors = inspect.get("errors", [])
            QMessageBox.warning(
                self,
                "리소스 확인",
                "작업 폴더의 resources 구성이 올바르지 않습니다.\n\n" + "\n".join(errors[:10])
            )
            return

        self.worker_name = worker_text
        self.work_root = work_root_path

        # 명단 파일 경로 및 열 매핑 — 미지정 시 진행 차단
        roster_log_text = self.setup_page.roster_log_input.text().strip()
        if not roster_log_text:
            QMessageBox.warning(
                self, "명단 파일 필요",
                "학교 전체 명단 파일(.xlsx)을 지정해야 합니다.\n"
                "'찾아보기' 버튼으로 파일을 선택하고 열 매핑을 완료해 주세요."
            )
            return
        roster_log_path_candidate = Path(roster_log_text)
        if not roster_log_path_candidate.exists():
            QMessageBox.warning(
                self, "명단 파일 없음",
                f"지정된 명단 파일을 찾을 수 없습니다.\n{roster_log_text}\n\n"
                "파일 경로를 다시 확인하거나 '찾아보기'로 재선택해 주세요."
            )
            return

        cfg = load_app_config()
        col_map_candidate = cfg.get("roster_col_map", {})
        if not col_map_candidate.get("col_school"):
            QMessageBox.warning(
                self, "열 매핑 필요",
                "명단 파일의 열 매핑이 완료되지 않았습니다.\n"
                "'찾아보기' 버튼으로 파일을 선택하고 열 매핑을 완료해 주세요."
            )
            return
        if not col_map_candidate.get("col_domain"):
            QMessageBox.warning(
                self, "열 매핑 필요",
                "명단 파일에서 도메인(홈페이지) 열이 지정되지 않았습니다.\n"
                "'찾아보기' 버튼으로 파일을 선택하고 열 매핑을 완료해 주세요."
            )
            return

        self.roster_log_path = roster_log_path_candidate
        self._roster_col_map = col_map_candidate

        qd = self.setup_page.work_date_edit.date()
        self.work_date = _date(qd.year(), qd.month(), qd.day())

        qd = self.setup_page.open_date_edit.date()
        self.school_start_date = _date(qd.year(), qd.month(), qd.day())
        self._school_year = self.school_start_date.year

        work_date_str = self.work_date.strftime("%Y-%m-%d")
        self.main_page.status_panel.current_worker.setText(f"작업자 · {self.worker_name}")
        self.main_page.status_panel.current_work_date.setText(f"작업일 · {work_date_str}")

        # 발송일 기본값 = 작업일
        work_qdate = QDate(self.work_date.year, self.work_date.month, self.work_date.day)
        self.main_page.status_panel.email_sent_date.setDate(work_qdate)
        self.main_page.header.badge_worker.setText(f"작업자 {self.worker_name}")
        self.main_page.step_bar.set_state(0, "done")

        self.setup_page.stat_last_date.set_value(work_date_str)
        self.setup_page.stat_last_step.set_value("기본 설정 완료")

        self.main_page.status_panel.btn_select_school.setEnabled(False)
        self.main_page.main_tab.btn_scan.setEnabled(False)
        self.main_page.main_tab.btn_run.setEnabled(False)
        self.main_page.main_tab.btn_run_diff.setEnabled(False)
        self.main_page.status_panel.school_status_label.setText("학교명을 입력해 검색하세요.")

        self._load_school_list()
        self.stack.setCurrentIndex(1)
        self.main_page.status_panel.school_input.setFocus()


    def apply_selected_school(self):
        from PyQt6.QtWidgets import QMessageBox

        # 타이머가 대기 중이면 즉시 flush해서 버튼 enable 상태를 최신으로 맞춤
        if self._school_search_timer.isActive():
            self._school_search_timer.stop()
            self._apply_school_search()

        # 명단 미기록 상태면 먼저 확인
        if not self._check_pending_roster_before("학교 변경"):
            return

        name = self.main_page.status_panel.school_input.text().strip()

        if not name:
            self.main_page.status_panel.school_status_label.setText("학교를 입력하세요.")
            QMessageBox.warning(self, "학교 선택", "학교를 입력하세요.")
            return

        # 1) DB exact match 확인
        if name not in self._school_name_set:
            self.selected_school = None
            self.main_page.status_panel.btn_select_school.setEnabled(False)
            self.main_page.status_panel.school_status_label.setText("DB에 없는 학교입니다.")
            QMessageBox.warning(self, "학교 선택", "DB에 없는 학교입니다.")
            return

        # 2) 실제 학교 폴더 매칭
        school_dirs = [
            p for p in Path(self.work_root).iterdir()
            if p.is_dir()
            and "resources" not in p.name.lower()
            and not p.name.startswith(".")
        ]

        matched = [p for p in school_dirs if name in p.name]

        if not matched:
            self.selected_school = None
            self.main_page.status_panel.school_status_label.setText("DB에는 있지만 작업 폴더 안에 학교 폴더가 없습니다.")
            QMessageBox.warning(
                self,
                "학교 폴더 없음",
                "DB에는 등록되어 있지만 작업 폴더 안에 해당 학교 폴더가 없습니다."
            )
            return

        if len(matched) > 1:
            self.selected_school = None
            self.main_page.status_panel.school_status_label.setText("일치하는 학교 폴더가 여러 개입니다.")
            QMessageBox.warning(
                self,
                "학교 폴더 중복",
                "해당 학교명이 포함된 폴더가 여러 개입니다.\n"
                + "\n".join(p.name for p in matched)
            )
            return

        # 3) 통과
        school = name
        self.selected_school = school

        btn = self.main_page.status_panel.btn_select_school
        btn.setText("적용됨")
        btn.setEnabled(False)
        btn.setStyleSheet("""
            QPushButton {
                background: #DCFCE7;
                border: 1px solid #16A34A;
                border-radius: 10px;
                padding: 10px 12px;
                font-weight: 700;
                color: #15803D;
                min-height: 22px;
            }
        """)

        self.main_page.status_panel.current_school.setText(school)
        self.main_page.status_panel.set_grade_count(school)
        self.main_page.status_panel.school_status_label.setText("")
        self.main_page.status_panel.school_status_label.hide()

        # 작업 이력 표시
        school_year = getattr(self, "_school_year", None) or QDate.currentDate().year()
        history = load_work_history(school_year)
        entry = history.get(school)
        lbl = self.main_page.status_panel.school_history_label
        if entry:
            last_date = entry.get("last_date", "")
            counts = entry.get("counts", {})
            count_str = " · ".join(f"{k} {v}명" for k, v in counts.items() if v)
            lbl.setText(f"마지막 작업 · {last_date}" + (f"\n{count_str}" if count_str else ""))
        else:
            lbl.setText("작업 이력 없음")
        lbl.show()

        self.main_page.step_bar.set_state(1, "done")
        self.main_page.step_bar.set_state(2, "active")
        self.main_page.status_panel.latest_log_label.setText(
            f"순번 12 · 작업자 {self.worker_name or '-'} · 상태 완료"
        )
        self.main_page.header.school_name.setText(f"현재 학교 · {school}")
        self.main_page.header.badge_status.setText("상태 학교 선택 완료")

        # 다음 단계 열기
        self.main_page.main_tab.btn_scan.setEnabled(True)
        self.main_page.main_tab.btn_run.setEnabled(False)
        self.main_page.main_tab.btn_run_diff.setEnabled(True)

        self.main_page.status_panel.school_result_list.hide()

        self._reset_scan_run_state()

        # 스캔 탭으로 자동 이동
        self.main_page.main_tab.tabs.setCurrentIndex(0)

        self._selected_domain = get_school_domain(
            self.roster_log_path,
            school,
            self._roster_col_map,
        ) or ""
        self._refresh_notice_tab()


    def handle_school_input_keypress(self, event):
        from PyQt6.QtCore import Qt

        school_input = self.main_page.status_panel.school_input
        school_list = self.main_page.status_panel.school_result_list

        key = event.key()

        if key in (Qt.Key.Key_Down, Qt.Key.Key_Up):
            if school_list.isHidden() or school_list.count() == 0:
                return False

            current_row = school_list.currentRow()
            if current_row < 0:
                current_row = 0 if key == Qt.Key.Key_Down else school_list.count() - 1
            else:
                if key == Qt.Key.Key_Down:
                    current_row = min(current_row + 1, school_list.count() - 1)
                else:
                    current_row = max(current_row - 1, 0)

            school_list.setCurrentRow(current_row)
            item = school_list.currentItem()
            if item:
                school_input.setText(item.text())
                school_input.selectAll()
                self.validate_school_input(item.text())
            return True

        if key in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            # 1순위: 리스트 선택 항목
            item = school_list.currentItem()
            if item is not None and not school_list.isHidden():
                text = item.text().strip()
                school_input.setText(text)
                self.validate_school_input(text)
                if text in self._school_name_set:
                    self.apply_selected_school()
                    return True

            # 2순위: 입력창 exact match
            text = school_input.text().strip()
            self.validate_school_input(text)
            if text in self._school_name_set:
                self.apply_selected_school()
                return True

            return True

        return False

    def sync_task_name(self):
        # diff 모드면 content_stack index가 1
        if self.main_page.content_stack.currentIndex() == 1:
            self.main_page.header.badge_task.setText("작업 재학생 명단 비교")
        else:
            idx = self.main_page.main_tab.tabs.currentIndex()
            _STEP_NAMES = {0: "스캔 검수", 1: "실행·결과", 2: "안내문"}
            task_name = _STEP_NAMES.get(idx, "")
            self.main_page.header.badge_task.setText(f"작업 {task_name}")
            
    def mark_scan_done(self):
        self.main_page.step_bar.set_state(2, "done")
        self.main_page.step_bar.set_state(3, "active")
        self.main_page.main_tab.scan_status_label.setText("스캔 완료")
        self.main_page.main_tab.scan_status_label.setStyleSheet(_STATUS_STYLE_OK)
        self.main_page.header.badge_status.setText("상태 스캔 완료")
        self.main_page.main_tab.btn_run.setEnabled(True)
        # btn_goto_run_tab은 _reveal_scan_result에서 활성화

    def mark_run_done(self, result=None):
        self.main_page.step_bar.set_state(3, "done")
        self.main_page.step_bar.set_state(4, "active")
        self.main_page.header.badge_status.setText("상태 실행 완료")

        # 실행 결과를 이력용으로 저장해두되 표시는 명단 기록 완료 시점에
        if self.selected_school and result is not None and result.ok:
            from datetime import date as _date_cls
            ic = (getattr(result, "audit_summary", None) or {}).get("input_counts", {})
            counts = {}
            if ic.get("freshmen", 0): counts["신입생"] = ic["freshmen"]
            if ic.get("transfer", 0): counts["전입"] = ic["transfer"]
            if ic.get("withdraw", 0): counts["전출"] = ic["withdraw"]
            if ic.get("teacher", 0): counts["교직원"] = ic["teacher"]
            self._pending_history_entry = {
                "last_date": _date_cls.today().strftime("%Y-%m-%d"),
                "worker": self.worker_name or "",
                "counts": counts,
            }
        else:
            self._pending_history_entry = None

        hold_count = 0
        auto_skip_count = 0
        if result is not None:
            hold_count = (
                getattr(result, "transfer_in_hold", 0)
                + getattr(result, "transfer_out_hold", 0)
            )
            auto_skip_count = getattr(result, "transfer_out_auto_skip", 0)

        real_hold = hold_count - auto_skip_count  # 자동제외 제외한 실제 보류

        if real_hold > 0 and auto_skip_count > 0:
            msg = f"보류 {real_hold}건이 있습니다. 생성된 파일의 보류 시트를 확인해 주세요. (자동제외 {auto_skip_count}건 별도)"
        elif real_hold > 0:
            msg = f"보류 {real_hold}건이 있습니다. 생성된 파일의 보류 시트를 확인해 주세요."
        else:
            msg = None  # 자동제외만 있는 경우는 run_hold_warning에 표시 안 함

        if msg:
            self.main_page.main_tab.run_hold_warning.setText(msg)
            self.main_page.main_tab.run_hold_warning.show()
        else:
            self.main_page.main_tab.run_hold_warning.hide()

        # 상태 요약에도 자동제외 반영
        if auto_skip_count > 0:
            self.main_page.main_tab.sum_withdraw.setText(
                self.main_page.main_tab.sum_withdraw.text()
                + f" (자동제외 {auto_skip_count}건)"
            )

        # 명단 기록 대기 상태로 전환
        if self.roster_log_path:
            self._pending_roster_log = True
            self.main_page.main_tab.btn_record_roster.setEnabled(True)
            self.main_page.status_panel.btn_record_roster_sidebar.setEnabled(True)

    def run_scan(self):
        if not self.work_root:
            self.main_page.main_tab.scan_message.setText("작업 폴더가 설정되지 않았습니다.")
            return
        if not self.selected_school:
            self.main_page.main_tab.scan_message.setText("학교를 먼저 선택해 주세요.")
            return

        tab = self.main_page.main_tab
        tab.btn_scan.setEnabled(False)   # 중복 실행 방지
        tab.scan_message.setText("스캔 중...")
        tab.scan_status_label.setText("스캔 중")
        tab.scan_status_label.setStyleSheet(_STATUS_STYLE_RUNNING)

        self._scan_thread = QThread(self)
        self._scan_worker = ScanWorker(
            work_root=self.work_root,
            school_name=self.selected_school,
            school_start_date=self.school_start_date,
            work_date=self.work_date,
            roster_xlsx=self.roster_log_path,
            col_map=self._roster_col_map,
        )
        self._scan_worker.moveToThread(self._scan_thread)

        self._scan_thread.started.connect(self._scan_worker.run)
        self._scan_worker.finished.connect(self._on_scan_finished)
        self._scan_worker.failed.connect(self._on_scan_failed)

        self._scan_worker.finished.connect(self._scan_thread.quit)
        self._scan_worker.failed.connect(self._scan_thread.quit)

        self._scan_worker.finished.connect(self._scan_worker.deleteLater)
        self._scan_worker.failed.connect(self._scan_worker.deleteLater)
        self._scan_thread.finished.connect(self._scan_thread.deleteLater)

        self._scan_thread.start()



    def _on_scan_finished(self, result):
        self.main_page.main_tab.btn_scan.setEnabled(True)
        tab = self.main_page.main_tab

        if not result.ok:
            # [ERROR] — 실행 중단. 빨간 상태 라벨 + 팝업
            self._last_scan_logs = result.logs or []
            self._last_scan_result = None
            first_error = next((l for l in result.logs if "[ERROR]" in l), None)
            msg = re.sub(r"^\[ERROR\]\s*", "", first_error or "스캔 중 오류가 발생했습니다.")
            tab.scan_message.setText(msg)
            tab.scan_status_label.setText("스캔 실패")
            tab.scan_status_label.setStyleSheet(_STATUS_STYLE_ERROR)
            tab.school_kind_warn_label.hide()
            tab.school_kind_row_widget.hide()
            self._school_kind_active = False
            self._school_kind_override = None
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "스캔 실패",
                "파일 스캔 중 오류가 발생했습니다.\n'스캔 로그 보기'에서 자세한 내용을 확인해 주세요.")
            return

        self._last_scan_result = result
        self._last_scan_logs = result.logs or []
        self.mark_scan_done()
        self._apply_scan_result(result)
        self._update_grade_map_from_scan(result)

        # 학교 구분 자동 판별 실패 여부 감지
        kind_warn = any("학교 구분을 자동으로 판별하지 못했습니다" in l for l in result.logs)
        tab.school_kind_warn_label.setVisible(kind_warn)
        tab.school_kind_row_widget.setVisible(kind_warn)
        self._school_kind_active = kind_warn
        if not kind_warn:
            self._school_kind_override = None

        # [WARN] — 노란 상태 라벨 + 경고 배너 (팝업 없음)
        warn_logs = [l for l in result.logs if "[WARN]" in l and "[DEBUG]" not in l]
        if warn_logs:
            first_warn = re.sub(r"^\[WARN\]\s*", "", warn_logs[0])
            tab.scan_status_label.setText("경고")
            tab.scan_status_label.setStyleSheet(_STATUS_STYLE_WARN)
            # 명부 없음 경고는 별도 안내문으로
            if result.need_roster and not result.roster_path:
                tab.scan_message.setText(
                    "학생명부를 찾지 못했습니다. "
                    "학년도 아이디 규칙을 직접 입력하거나 명부를 추가한 뒤 재스캔해 주세요."
                )
            else:
                tab.scan_message.setText(f"경고 {len(warn_logs)}건 — {first_warn}")
        else:
            tab.scan_status_label.setText("스캔 완료")
            tab.scan_status_label.setStyleSheet(_STATUS_STYLE_OK)
            tab.scan_message.setText("스캔 완료 — 이상 없음")

        # 스캔 결과 요약 — 스캔 표에 올라온 파일 구분 목록으로 구성
        present = []
        kind_order = ["신입생", "전입생", "전출생", "교직원"]
        for kind in kind_order:
            if kind in tab.scan_preview_data:
                present.append(kind)
        summary = "  ·  ".join(present) + " 파일 검수 완료" if present else "파일 검수 완료"
        if warn_logs:
            summary += f"  (경고 {len(warn_logs)}건)"
        self._reveal_scan_result(summary)

    def _on_scan_failed(self, error_text):
        # Worker에서 잡히지 않은 예상치 못한 예외 (traceback 포함)
        self.main_page.main_tab.btn_scan.setEnabled(True)
        self._last_scan_result = None
        self._last_scan_logs = [f"[DEBUG] {error_text}"]
        self.main_page.main_tab.scan_message.setText("예기치 못한 오류가 발생했습니다.")
        self.main_page.main_tab.scan_status_label.setText("스캔 실패")
        self.main_page.main_tab.scan_status_label.setStyleSheet(_STATUS_STYLE_ERROR)
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.warning(self, "스캔 오류",
            "예기치 못한 오류가 발생했습니다.\n\n자세한 내용은 '스캔 로그 보기'를 확인해 주세요.")

    def _collect_layout_overrides(self) -> dict:
        """
        스캔 표의 '수정 시작행' 스핀 위젯 값과
        학년도 아이디 규칙 수동 설정값을 읽어
        execute_pipeline에 넘길 layout_overrides dict를 구성한다.
        """
        table = self.main_page.main_tab.scan_table
        kind_key_map = {"신입생": "freshmen", "전입생": "transfer",
                        "전출생": "withdraw", "교직원": "teacher"}
        overrides = {}
        for row, kind_key in [
            (0, "freshmen"), (1, "transfer"), (2, "withdraw"), (3, "teacher")
        ]:
            kind_label = next((k for k, v in kind_key_map.items() if v == kind_key), None)
            if kind_label is None:
                continue
            row_idx = self._SCAN_TABLE_KIND_ROW.get(kind_label)
            if row_idx is None:
                continue
            spin_widget = table.cellWidget(row_idx, 4)
            if spin_widget is None:
                continue
            label = spin_widget.findChild(QLabel)
            if label is None:
                continue
            try:
                val = int(label.text())
                if val > 0:
                    overrides[kind_key] = {"data_start_row": val}
            except ValueError:
                pass

        # 학년도 아이디 규칙 수동 오버라이드
        grade_years = self.main_page.status_panel.get_grade_year_overrides()
        if grade_years:
            overrides["grade_year_map"] = grade_years

        return overrides

    def _on_grade_map_apply(self):
        """학년도 아이디 규칙 수동 적용 확인 메시지."""
        grade_years = self.main_page.status_panel.get_grade_year_overrides()
        if not grade_years:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "학년도 아이디 규칙", "입력된 학년도 값이 없습니다.")
            return
        lines = [f"{g}학년 → {y}" for g, y in sorted(grade_years.items())]
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.information(
            self, "학년도 아이디 규칙 적용",
            "다음 값으로 실행 시 적용됩니다:\n\n" + "\n".join(lines)
        )

    def run_main(self):
        if not self.work_root:
            self.main_page.main_tab.run_info.setText("작업 폴더가 설정되지 않았습니다.")
            return
        if not self.selected_school:
            self.main_page.main_tab.run_info.setText("학교를 먼저 선택해 주세요.")
            return
        if not self._last_scan_result:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "실행 불가", "스캔을 먼저 완료해야 합니다.")
            return

        # 체크박스 확인은 → 버튼 클릭 시(_on_goto_run_tab)에서 처리

        tab = self.main_page.main_tab
        tab.btn_run.setEnabled(False)
        tab.run_info.setText("실행 중...")
        tab.run_hold_warning.hide()
        tab.run_status_label.setText("실행 중")
        tab.run_status_label.setStyleSheet(_STATUS_STYLE_RUNNING)

        layout_overrides = self._collect_layout_overrides()

        # 학교 구분 자동 판별 실패 시 사용자가 선택한 값 수집
        school_kind_override = None
        if getattr(self, "_school_kind_active", False):
            school_kind_override = tab.school_kind_combo.currentText() or None

        self._run_thread = QThread(self)
        self._run_worker = RunWorker(
            scan=self._last_scan_result,
            work_date=self.work_date,
            school_start_date=self.school_start_date,
            layout_overrides=layout_overrides,
            school_kind_override=school_kind_override,
        )
        self._run_worker.moveToThread(self._run_thread)

        self._run_thread.started.connect(self._run_worker.run)
        self._run_worker.finished.connect(self._on_run_finished)
        self._run_worker.failed.connect(self._on_run_failed)

        self._run_worker.finished.connect(self._run_thread.quit)
        self._run_worker.failed.connect(self._run_thread.quit)

        self._run_worker.finished.connect(self._run_worker.deleteLater)
        self._run_worker.failed.connect(self._run_worker.deleteLater)
        self._run_thread.finished.connect(self._run_thread.deleteLater)

        self._run_thread.start()

    def _on_run_finished(self, result):
        self.main_page.main_tab.btn_run.setEnabled(True)
        self._last_run_logs = result.logs or []
        tab = self.main_page.main_tab

        if not result.ok:
            first_error = next((l for l in result.logs if "[ERROR]" in l), None)
            msg = re.sub(r"^\[ERROR\]\s*", "", first_error or "실행 중 오류가 발생했습니다.")
            tab.run_info.setText(msg)
            tab.run_status_label.setText("실행 실패")
            tab.run_status_label.setStyleSheet(_STATUS_STYLE_ERROR)
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "실행 실패",
                "작업 실행 중 오류가 발생했습니다.\n'실행 로그 보기'에서 자세한 내용을 확인해 주세요.")
            return

        self._current_output_files = list(result.outputs or [])
        self._last_run_result = result
        self.mark_run_done(result)
        self._apply_run_result(result)

        # [WARN] — 노란 상태 라벨 (팝업 없음, 계속 진행)
        warn_logs = [l for l in result.logs if "[WARN]" in l and "[DEBUG]" not in l]
        if warn_logs:
            first_warn = re.sub(r"^\[WARN\]\s*", "", warn_logs[0])
            tab.run_status_label.setText("경고")
            tab.run_status_label.setStyleSheet(_STATUS_STYLE_WARN)
            tab.run_info.setText(f"실행 완료 — 경고 {len(warn_logs)}건: {first_warn}")
        else:
            tab.run_status_label.setText("실행 완료")
            tab.run_status_label.setStyleSheet(_STATUS_STYLE_OK)
            tab.run_info.setText("실행 완료")

        # 실행 결과 요약 텍스트 조합 후 하단 결과 영역 펼침
        parts = []
        freshmen = getattr(result, "freshmen_count", None)
        transfer = getattr(result, "transfer_in_done", None)
        withdraw = getattr(result, "transfer_out_done", None)
        teacher  = getattr(result, "teacher_count", None)
        hold     = getattr(result, "transfer_in_hold", 0) + getattr(result, "transfer_out_hold", 0)
        auto_skip = getattr(result, "transfer_out_auto_skip", 0)
        real_hold_count = hold - auto_skip
        if freshmen is not None: parts.append(f"신입생 {freshmen}명")
        if transfer  is not None: parts.append(f"전입 {transfer}건")
        if withdraw  is not None: parts.append(f"전출 {withdraw}건")
        if teacher   is not None: parts.append(f"교직원 {teacher}명")
        if real_hold_count > 0:   parts.append(f"보류 {real_hold_count}건")
        if warn_logs:             parts.append(f"경고 {len(warn_logs)}건")
        run_summary = "  ·  ".join(parts) if parts else "작업 완료"
        self._reveal_run_result(run_summary)

    def _on_run_failed(self, error_text):
        # Worker에서 잡히지 않은 예상치 못한 예외 (traceback 포함)
        self.main_page.main_tab.btn_run.setEnabled(True)
        self._last_run_logs = [f"[DEBUG] {error_text}"]
        tab = self.main_page.main_tab
        tab.run_info.setText("예기치 못한 오류가 발생했습니다.")
        tab.run_status_label.setText("실행 실패")
        tab.run_status_label.setStyleSheet(_STATUS_STYLE_ERROR)
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.warning(self, "실행 오류",
            "예기치 못한 오류가 발생했습니다.\n\n자세한 내용은 '실행 로그 보기'를 확인해 주세요.")

    # ------------------------------------------------------------------
    # 결과 → UI 변환 헬퍼
    # ------------------------------------------------------------------

    def _safe_str(self, value):
        if value is None:
            return ""
        return str(value)

    def _rows_to_table_rows(self, rows, headers=None, limit=None):
        out = []
        if rows is None:
            return out
        for row in (rows[:limit] if limit else rows):
            if isinstance(row, dict):
                if headers:
                    out.append([self._safe_str(row.get(h, "")) for h in headers])
                else:
                    out.append([self._safe_str(v) for v in row.values()])
            elif isinstance(row, (list, tuple)):
                out.append([self._safe_str(v) for v in row])
            else:
                out.append([self._safe_str(row)])
        return out

    def _extract_scan_items(self, result):
        """
        ScanResult에서 파일별 스캔 메타를 유연하게 꺼낸다.
        vars()와 known-key 탐색을 합쳐 중복 없이 수집.
        """
        KNOWN_KEYS = ["freshmen", "transfer_in", "transfer_out", "teachers", "roster"]
        seen = set()
        candidates = []

        # known key를 우선 수집 (라벨 매핑 보장)
        for key in KNOWN_KEYS:
            value = getattr(result, key, None)
            if isinstance(value, dict) and {"file_name", "sheet_name"} & set(value.keys()):
                seen.add(key)
                candidates.append((key, value))

        # known key에 없는 나머지 dict 필드도 탐색
        for key, value in vars(result).items():
            if key in seen:
                continue
            if isinstance(value, dict) and {"file_name", "sheet_name"} & set(value.keys()):
                seen.add(key)
                candidates.append((key, value))

        return candidates

    def _apply_scan_result(self, result):
        tab = self.main_page.main_tab
        tab.scan_preview_data = {}

        items = self._extract_scan_items(result)

        # 스캔 표에 올릴 항목 (학생명부 제외 — 시작행 조정 불필요)
        TABLE_KINDS = {"신입생", "전입생", "전출생", "교직원"}

        kind_label_map = {
            "roster":      "학생명부",
            "freshmen":    "신입생",
            "transfer_in": "전입생",
            "transfer_out":"전출생",
            "teachers":    "교직원",
        }

        table_items = []   # 스캔 표 갱신용
        has_preview = False

        for raw_kind, item in items:
            kind = kind_label_map.get(raw_kind, raw_kind)

            file_name      = self._safe_str(item.get("file_name", ""))
            sheet_name     = self._safe_str(item.get("sheet_name", ""))
            header_row     = item.get("header_row", "")
            data_start_row = item.get("data_start_row", "")
            warning        = self._safe_str(item.get("warning", ""))
            headers        = item.get("headers") or item.get("columns") or []
            rows           = item.get("rows") or item.get("preview_rows") or []

            # 미리보기 데이터는 학생명부 포함 전부 저장
            tab.scan_preview_data[kind] = {
                "file_name":     file_name,
                "sheet_name":    sheet_name,
                "header_row":    header_row,
                "data_start_row":data_start_row,
                "warning":       warning or "이상 없음",
                "headers":       headers,
                "rows":          self._rows_to_table_rows(rows, headers=headers),
                "issue_rows":    set(item.get("issue_rows", [])),
            }
            has_preview = True

            # 스캔 표는 TABLE_KINDS만
            if kind in TABLE_KINDS:
                table_items.append({
                    "kind":          kind,
                    "file_name":     file_name,
                    "sheet_name":    sheet_name,
                    "data_start_row":data_start_row,
                })

        self._update_scan_table_from_items(table_items)

        # 파일 없는 구분의 스핀/체크 비활성화
        present_kinds = {item["kind"] for item in table_items}
        table = self.main_page.main_tab.scan_table
        for kind, row in self._SCAN_TABLE_KIND_ROW.items():
            has_file = kind in present_kinds
            spin_widget = table.cellWidget(row, 4)
            if spin_widget is not None:
                spin_widget.setEnabled(has_file)
                for c in range(1, 4):
                    cell = table.item(row, c)
                    if cell:
                        cell.setText("" if not has_file else cell.text())
            chk_widget = table.cellWidget(row, 5)
            if chk_widget is not None:
                chk = chk_widget.findChild(QCheckBox)
                if chk is not None:
                    if not has_file:
                        chk.setChecked(False)
                    chk_widget.setEnabled(has_file)

        if has_preview:
            # Bug 2 수정: selectRow가 itemSelectionChanged 시그널을 발생시켜
            # 빈 상태에서 on_scan_table_row_selected가 먼저 호출되는 문제 차단
            tab.scan_table.blockSignals(True)
            tab.scan_table.selectRow(0)
            tab.scan_table.blockSignals(False)

            first_kind = next(
                (k for k in ["신입생", "전입생", "전출생", "교직원"] if k in tab.scan_preview_data),
                None
            )
            if first_kind:
                tab.load_selected_preview(first_kind)

            # 스캔 완료 시 뷰어 항상 자동 펼침
            tab.btn_toggle_viewer.setChecked(True)
            tab.viewer_body.setVisible(True)
            tab.btn_toggle_viewer.setText("접기 ▴")
        else:
            tab.preview_file_info.setText("파일: - | 시트: - | 헤더행: - | 시작행: -")
            tab.preview_warning_label.setText("표시할 스캔 미리보기가 없습니다.")

    # 스캔 표 고정 행 순서
    _SCAN_TABLE_KIND_ROW = {
        "신입생": 0,
        "전입생": 1,
        "전출생": 2,
        "교직원":   3,
    }

    def _update_scan_table_from_items(self, items):
        """
        기존 4행 편집형 스캔 표의 텍스트만 갱신.
        컬럼 구조·셀 위젯(+/- 스핀, 체크박스)은 절대 건드리지 않는다.
        """
        table = self.main_page.main_tab.scan_table

        for item in items:
            kind = item.get("kind", "")
            row = self._SCAN_TABLE_KIND_ROW.get(kind)
            if row is None:
                continue

            # 0~3열: 구분 / 파일명 / 시트 / 자동 감지(data_start_row)
            text_values = [
                kind,
                item.get("file_name", ""),
                item.get("sheet_name", ""),
                self._safe_str(item.get("data_start_row", "")),
            ]
            for c, value in enumerate(text_values):
                cell = table.item(row, c)
                if cell is None:
                    cell = QTableWidgetItem()
                    cell.setFlags(cell.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    table.setItem(row, c, cell)
                cell.setText(value)

            # 4열: 수정 시작행 위젯 내부 숫자 라벨을 감지값으로 세팅
            spin_widget = table.cellWidget(row, 4)
            if spin_widget is not None:
                label = spin_widget.findChild(QLabel)
                if label is not None:
                    detected = self._safe_str(item.get("data_start_row", ""))
                    if detected:
                        label.setText(detected)

            # 5열: 체크박스 초기화 (새 스캔마다 unchecked)
            chk_widget = table.cellWidget(row, 5)
            if chk_widget is not None:
                chk = chk_widget.findChild(QCheckBox)
                if chk is not None:
                    chk.setChecked(False)

    # ------------------------------------------------------------------
    # 실행 결과 → UI
    # ------------------------------------------------------------------

    def _extract_output_paths(self, result):
        outputs = getattr(result, "outputs", None) or []
        out = []
        for value in outputs:
            try:
                out.append(str(value))
            except Exception:
                pass
        return out

    def _build_run_summary_from_result(self, result):
        audit = getattr(result, "audit_summary", None) or {}
        input_counts  = audit.get("input_counts", {})
        result_counts = audit.get("result_counts", {})
        checks        = audit.get("checks", {})

        return {
            "school_name": audit.get("school_name") or self.selected_school or "",
            "year_str":    audit.get("year_str") or self._safe_str(getattr(result, "year_str", "")),
            "input_counts": {
                "freshmen": input_counts.get("freshmen", getattr(result, "freshmen_count", "")),
                "teacher":  input_counts.get("teacher",  getattr(result, "teacher_count", "")),
                "transfer": input_counts.get("transfer", getattr(result, "transfer_in_count", "")),
                "withdraw": input_counts.get("withdraw", getattr(result, "transfer_out_count", "")),
            },
            "result_counts": {
                "transfer_done":    result_counts.get("transfer_done",    getattr(result, "transfer_in_done", "")),
                "transfer_hold":    result_counts.get("transfer_hold",    getattr(result, "transfer_in_hold", "")),
                "withdraw_done":    result_counts.get("withdraw_done",    getattr(result, "transfer_out_done", "")),
                "withdraw_hold":    result_counts.get("withdraw_hold",    getattr(result, "transfer_out_hold", "")),
                "withdraw_auto_skip": result_counts.get("withdraw_auto_skip", getattr(result, "transfer_out_auto_skip", "")),
            },
            "checks": {
                "transfer_total_match": checks.get("transfer_total_match", None),
                "withdraw_total_match": checks.get("withdraw_total_match", None),
            },
        }

    def _detect_output_header_row(self, file_path: Path) -> int:
        """파일명 기준으로 헤더 행 번호를 반환 (1-indexed)."""
        name = Path(file_path).name
        if "안내" in name:
            return 3
        return 1  # 등록 파일 및 기타

    def _read_xlsx_sheets(self, file_path, max_rows: int = 5000) -> dict:
        """
        xlsx 파일을 열어 시트별 headers + rows를 반환한다.
        헤더 행은 파일명 규칙으로 결정 (탐지 없음).
        실패 시 에러 시트 하나로 fallback.
        """
        try:
            from openpyxl import load_workbook
            file_path = Path(file_path)
            header_row = self._detect_output_header_row(file_path)

            wb = load_workbook(file_path, data_only=True)
            sheets = {}

            for ws in wb.worksheets:
                # 관리자 전용 시트는 파일 뷰어에서 제외 (사용자에게 불필요)
                if "관리자" in ws.title:
                    continue

                max_col = ws.max_column or 1
                max_row = ws.max_row or 1

                # 헤더
                headers = []
                for c in range(1, max_col + 1):
                    val = ws.cell(row=header_row, column=c).value
                    headers.append("" if val is None else str(val))

                # 데이터: 헤더 다음 행부터, 완전 빈 행 제외, max_rows 제한
                data_start = header_row + 1
                rows = []
                for r in range(data_start, min(max_row + 1, data_start + max_rows)):
                    row_vals = [
                        "" if ws.cell(row=r, column=c).value is None
                        else str(ws.cell(row=r, column=c).value)
                        for c in range(1, max_col + 1)
                    ]
                    if any(v != "" for v in row_vals):
                        rows.append(row_vals)

                sheets[ws.title] = {"headers": headers, "rows": rows}

            wb.close()
            return sheets

        except Exception as e:
            return {"(읽기 오류)": {"headers": ["오류"], "rows": [[str(e)]]}}

    def _apply_run_result(self, result):
        tab = self.main_page.main_tab
        tab.run_preview_data = {}

        output_paths = self._extract_output_paths(result)
        summary = self._build_run_summary_from_result(result)
        log_rows = [[line] for line in (getattr(result, "logs", None) or ["실행 완료"])]
        log_sheet = {"로그": {"headers": ["로그"], "rows": log_rows}}

        if not output_paths:
            tab.run_preview_data["실행 결과"] = {
                "summary": summary,
                "sheets": {},
            }
        else:
            for output_path in output_paths:
                file_name = Path(output_path).name
                try:
                    real_sheets = self._read_xlsx_sheets(output_path)
                    sheets = real_sheets
                except Exception:
                    sheets = {}

                tab.run_preview_data[file_name] = {
                    "summary": summary,
                    "sheets": sheets,
                }

        # export_path에 저장 위치 표시
        if output_paths:
            tab.export_path.setText(str(Path(output_paths[0]).parent))
        else:
            tab.export_path.setText("")

        self._reload_run_preview_ui()

    def _reload_run_preview_ui(self):
        tab = self.main_page.main_tab

        file_names = list(tab.run_preview_data.keys())

        # 콤보박스 (하위 호환용)
        tab.run_file_combo.blockSignals(True)
        tab.run_file_combo.clear()
        tab.run_file_combo.addItems(file_names)
        tab.run_file_combo.blockSignals(False)

        # 파일명 표 채우기
        tbl = tab.run_file_list_widget
        tbl.setRowCount(0)
        for name in file_names:
            r = tbl.rowCount()
            tbl.insertRow(r)
            item = QTableWidgetItem(name)
            item.setToolTip("더블클릭하면 파일이 열립니다")
            tbl.setItem(r, 0, item)
        row_h = 32
        tbl.setFixedHeight(
            tbl.horizontalHeader().height() + max(1, len(file_names)) * row_h + 4
        )

        if file_names:
            tab.run_file_combo.setCurrentIndex(0)
            tab.load_selected_run_file(file_names[0])
        else:
            tab.current_run_file = None
            tab.run_preview_info.setText("시트: - | 행 수: -")

    def _update_grade_map_from_scan(self, result):
        panel = self.main_page.status_panel

        need_roster = getattr(result, "need_roster", False)
        if not need_roster:
            # 스캔 성공 + 명부 불필요 (신입생만 있는 작업 등)
            panel.update_grade_map(state="not_needed")
            return

        # 명부가 필요한데 roster_info가 없으면 읽지 못한 것
        roster_info = getattr(result, "roster_info", None)
        if not roster_info:
            panel.update_grade_map(state="no_roster")
            return

        # roster_info에서 학년별 prefix(기준 학년도) 추출
        prefix_by_grade = (
            roster_info.get("prefix_mode_by_roster_grade", {})
            if isinstance(roster_info, dict)
            else getattr(roster_info, "prefix_mode_by_roster_grade", {})
        ) or {}
        shift = int((
            roster_info.get("ref_grade_shift", 0)
            if isinstance(roster_info, dict)
            else getattr(roster_info, "ref_grade_shift", 0)
        ) or 0)

        mapping = {}
        for g_roster, prefix in prefix_by_grade.items():
            try:
                g_cur = int(g_roster) - shift
                if 1 <= g_cur <= 6:
                    mapping[g_cur] = int(prefix)
            except (TypeError, ValueError):
                continue

        # 1학년은 명부에 없는 경우가 많으므로 2학년 prefix +1 로 계산
        # (2학년도 없으면 3학년 기준으로 시도)
        if 1 not in mapping:
            for ref_grade in (2, 3):
                if ref_grade in mapping:
                    mapping[1] = mapping[ref_grade] + (ref_grade - 1)
                    break

        if mapping:
            panel.update_grade_map(mapping=mapping, state="ok")
        else:
            panel.update_grade_map(state="no_roster")

    def run_diff(self):
        tab = self.main_page.main_tab

        # ── guard ──────────────────────────────────────────────
        if not self.work_root:
            tab.diff_scan_info.setText("작업 폴더가 설정되지 않았습니다.")
            tab.diff_scan_info.show()
            return
        if not self.selected_school:
            tab.diff_scan_info.setText("학교를 먼저 선택해 주세요.")
            tab.diff_scan_info.show()
            return

        target_year = tab.diff_target_year.value()

        # ── 사전 점검 ───────────────────────────────────────────
        tab.diff_status_label.setText("점검 중...")
        tab.diff_status_label.setStyleSheet(_STATUS_STYLE_RUNNING)
        tab.diff_scan_info.hide()

        scan = scan_diff_engine(
            work_root=self.work_root,
            school_name=self.selected_school,
            target_year=target_year,
            school_start_date=self.school_start_date,
            work_date=self.work_date,
        )

        # 점검 결과 요약 구성
        scan_lines = []
        roster_ok  = bool(getattr(scan, "roster_path", None))
        compare_ok = bool(getattr(scan, "compare_file", None))
        scan_lines.append(f"기준 명부: {'인식 완료' if roster_ok else '인식 실패'}")
        scan_lines.append(f"비교 명단: {'인식 완료' if compare_ok else '인식 실패'}")

        missing = getattr(scan, "missing_fields", [])
        if missing:
            scan_lines.append(f"누락 항목: {', '.join(missing)}")

        shift = getattr(scan, "ref_grade_shift", 0)
        if shift:
            scan_lines.append(f"명부 기준 학년 차이: {shift:+d} (이전 학년도 명부로 간주)")

        tab.diff_scan_info.setText("\n".join(scan_lines))
        tab.diff_scan_info.show()

        if not scan.ok or not getattr(scan, "can_execute", False):
            reason = ", ".join(missing) if missing else "사전 점검 실패"
            self._reset_diff_result()
            tab.diff_summary.setText(f"실행 불가: {reason}")
            tab.diff_scan_info.show()
            tab.diff_status_label.setText("점검 실패")
            tab.diff_status_label.setStyleSheet(_STATUS_STYLE_ERROR)
            return

        # ── 실행 ───────────────────────────────────────────────
        tab.diff_status_label.setText("실행 중...")

        result = run_diff_engine(
            work_root=self.work_root,
            school_name=self.selected_school,
            target_year=target_year,
            school_start_date=self.school_start_date,
            work_date=self.work_date,
            roster_xlsx=self.roster_log_path,
            col_map=self._roster_col_map,
        )

        if not result.ok:
            self._last_diff_logs = result.logs or []
            err_msg = next((l for l in reversed(result.logs or []) if "[ERROR]" in l), "실행 실패")
            tab.diff_summary.setText(err_msg)
            tab.diff_status_label.setText("실행 실패")
            tab.diff_status_label.setStyleSheet(_STATUS_STYLE_ERROR)
            return

        # ── 성공 ───────────────────────────────────────────────
        self._last_diff_logs = result.logs or []
        self._current_output_files = list(result.outputs or [])
        tab.diff_status_label.setText("완료")
        tab.diff_status_label.setStyleSheet(_STATUS_STYLE_OK)

        # 요약 수치 4항목
        tab.diff_roster_only.setText(f"{result.roster_only_count}명")
        tab.diff_matched.setText(f"{result.matched_count}명")
        tab.diff_compare_only.setText(f"{result.compare_only_count}명")
        tab.diff_unresolved.setText(f"{result.unresolved_count}명")

        # 3열 집합 뷰어 채우기
        self._fill_diff_table(tab.diff_tbl_roster_only,  result.roster_only_rows,  max_rows=None)
        self._fill_diff_table(tab.diff_tbl_matched,      result.matched_rows,      max_rows=100)
        self._fill_diff_table(tab.diff_tbl_compare_only, result.compare_only_rows, max_rows=None)
        self._fill_diff_table(tab.diff_tbl_unresolved,   result.unresolved_rows,   max_rows=None,
                              has_reason=True)

        # 자동 분류 요약
        classify_lines = [
            f"자동 분류 전입  {result.transfer_in_done}명 / 확인 필요 {result.transfer_in_hold}명",
            f"자동 분류 전출  {result.transfer_out_done}명 / 확인 필요 {result.transfer_out_hold}명",
        ]
        tab.diff_summary.setText("\n".join(classify_lines))

        # 생성 파일 목록
        tab.diff_outputs.clear()
        for path in (result.outputs or []):
            tab.diff_outputs.addItem(str(Path(path).name))
        if not result.outputs:
            tab.diff_outputs.addItem("생성된 파일 없음")

        if result.outputs:
            tab.diff_output_dir.setText(str(Path(result.outputs[0]).parent))

    def _fill_diff_table(self, table: "QTableWidget", rows, max_rows=None, has_reason=False):
        """diff 결과 row 리스트를 QTableWidget에 채운다."""
        table.setRowCount(0)
        data = rows[:max_rows] if max_rows else rows
        for rec in data:
            r = table.rowCount()
            table.insertRow(r)
            table.setItem(r, 0, QTableWidgetItem(str(rec.get("grade", ""))))
            table.setItem(r, 1, QTableWidgetItem(str(rec.get("class", ""))))
            table.setItem(r, 2, QTableWidgetItem(str(rec.get("name", ""))))
            if has_reason:
                table.setItem(r, 3, QTableWidgetItem(str(rec.get("hold_reason", ""))))
        if max_rows and len(rows) > max_rows:
            r = table.rowCount()
            table.insertRow(r)
            remaining = len(rows) - max_rows
            item = QTableWidgetItem(f"... 외 {remaining}명")
            item.setForeground(QColor("#64748B"))
            table.setItem(r, 0, item)

    def _reset_diff_result(self):
        """학교 변경 또는 실패 시 diff 결과 영역 전체 초기화."""
        tab = self.main_page.main_tab
        tab.diff_roster_only.setText("-")
        tab.diff_matched.setText("-")
        tab.diff_compare_only.setText("-")
        tab.diff_unresolved.setText("-")
        for tbl in [tab.diff_tbl_roster_only, tab.diff_tbl_matched,
                    tab.diff_tbl_compare_only, tab.diff_tbl_unresolved]:
            tbl.setRowCount(0)
        tab.diff_summary.setText("명단 비교 실행 버튼을 눌러 주세요.")
        tab.diff_outputs.clear()
        tab.diff_output_dir.setText("")
        tab.diff_status_label.setText("실행 전")
        tab.diff_status_label.setStyleSheet(_STATUS_STYLE_IDLE)
        tab.diff_scan_info.hide()

    # ------------------------------------------------------------------
    # 로그 팝업
    # ------------------------------------------------------------------
    def _show_log_dialog(self, title: str, logs: list):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QPlainTextEdit, QDialogButtonBox

        if not logs:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, title, "표시할 로그가 없습니다.\n먼저 작업을 실행해 주세요.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle(title)
        dlg.resize(720, 520)

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        # [DEBUG] 줄은 로그 뷰어에 노출하지 않음 — 개발자용 내부 정보
        visible_logs = [l for l in logs if not l.startswith("[DEBUG]")]

        # 로그 레벨 정의:
        #   [ERROR] — 실행 중단이 필요한 사용자 설명 가능 오류
        #   [WARN]  — 진행은 됐지만 확인이 필요한 상황
        #   [INFO]  — 정상 진행 상황 보고
        #   [OK]    — 개별 단계(파일 인식, 데이터 로드 등) 성공
        #   [DONE]  — 파이프라인 전체 완료
        #   [DEBUG] — 개발자용 내부 상태 (로그 뷰어에 표시 안 함)

        text = QPlainTextEdit()
        text.setReadOnly(True)
        text.setPlainText("\n".join(visible_logs))
        text.setStyleSheet("""
            QPlainTextEdit {
                font-family: 'Consolas', 'D2Coding', monospace;
                font-size: 12px;
                background: #F8FAFC;
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                padding: 8px;
            }
        """)

        from PyQt6.QtGui import QTextCharFormat, QColor, QTextCursor
        cursor = text.textCursor()
        doc = text.document()
        block = doc.begin()
        while block.isValid():
            line = block.text()
            fmt = QTextCharFormat()
            if line.startswith("[DONE]"):
                fmt.setForeground(QColor("#15803D"))
            elif line.startswith("[WARN]"):
                fmt.setForeground(QColor("#B45309"))
            elif line.startswith("[ERROR]"):
                fmt.setForeground(QColor("#DC2626"))
            elif line.startswith("[INFO]"):
                fmt.setForeground(QColor("#1D4ED8"))
            else:
                fmt.setForeground(QColor("#334155"))
            cursor.setPosition(block.position())
            cursor.movePosition(QTextCursor.MoveOperation.EndOfBlock,
                                QTextCursor.MoveMode.KeepAnchor)
            cursor.setCharFormat(fmt)
            block = block.next()

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        btns.rejected.connect(dlg.accept)

        layout.addWidget(text)
        layout.addWidget(btns)
        dlg.exec()

    # ------------------------------------------------------------------
    # 파일 / 폴더 열기
    # ------------------------------------------------------------------
    def _on_run_file_double_clicked(self, item):
        """실행 결과 파일명 더블클릭 → 파일 열기"""
        import os
        file_name = item.text()
        # _current_output_files에서 파일명 매칭
        for path in self._current_output_files:
            if Path(path).name == file_name or str(path).endswith(file_name):
                try:
                    os.startfile(str(path))
                except Exception as e:
                    from PyQt6.QtWidgets import QMessageBox
                    QMessageBox.warning(self, "파일 열기 실패", str(e))
                return

    def _open_selected_file(self):
        import os
        tab = self.main_page.main_tab
        row = tab.run_file_combo.currentIndex()
        if row < 0 or row >= len(self._current_output_files):
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "파일 열기", "열 파일을 선택해 주세요.")
            return
        path = self._current_output_files[row]
        try:
            os.startfile(str(path))
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "오류", f"파일을 열 수 없습니다.\n{e}")

    def _open_scan_file_by_name(self, file_name: str):
        """스캔 표 더블클릭 → scan_preview_data에서 파일명으로 실제 경로를 찾아 열기."""
        import os
        if not file_name or not self._last_scan_result:
            return
        input_dir = getattr(self._last_scan_result, "input_dir", None)
        if not input_dir:
            return
        file_path = Path(input_dir) / file_name
        if not file_path.exists():
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "파일 열기", f"파일을 찾을 수 없습니다.\n{file_path}")
            return
        try:
            os.startfile(str(file_path))
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "파일 열기", f"파일을 열 수 없습니다.\n{e}")

    def _open_selected_scan_file(self):
        import os
        tab = self.main_page.main_tab
        row = tab.scan_table.currentRow()
        kind = ""
        if row >= 0:
            kind_item = tab.scan_table.item(row, 0)
            if kind_item:
                kind = kind_item.text().strip()

        # scan_preview_data에서 파일명 꺼내 실제 경로 찾기
        data = tab.scan_preview_data.get(kind, {})
        file_name = data.get("file_name", "")

        if not file_name or not self._last_scan_result:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "파일 열기", "열 파일을 선택해 주세요.")
            return

        # 실제 경로는 scan_result의 input_dir 기준
        input_dir = getattr(self._last_scan_result, "input_dir", None)
        if not input_dir:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "파일 열기", "파일 경로를 찾을 수 없습니다.")
            return

        file_path = Path(input_dir) / file_name
        if not file_path.exists():
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "파일 열기", f"파일을 찾을 수 없습니다.\n{file_path}")
            return

        try:
            os.startfile(str(file_path))
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "오류", f"파일을 열 수 없습니다.\n{e}")

    def _open_output_folder(self):
        import os
        # 출력 경로 우선: current_output_files → work_root
        folder = None
        if self._current_output_files:
            folder = self._current_output_files[0].parent
        elif self.work_root:
            folder = self.work_root
        if not folder:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "폴더 열기", "열 폴더 경로가 없습니다.\n먼저 작업을 실행해 주세요.")
            return
        try:
            os.startfile(str(folder))
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "오류", f"폴더를 열 수 없습니다.\n{e}")

    # ------------------------------------------------------------------
    # 찾아보기
    # ------------------------------------------------------------------
    def _browse_work_root(self):
        from PyQt6.QtWidgets import QFileDialog
        current = self.setup_page.work_root_input.text().strip()
        start = current if current else str(Path.home())
        folder = QFileDialog.getExistingDirectory(
            self, "작업 폴더 선택", start,
            QFileDialog.Option.ShowDirsOnly | QFileDialog.Option.DontResolveSymlinks
        )
        if folder:
            self.setup_page.work_root_input.setText(folder)
            self.setup_page._refresh_start_button()

    def _browse_roster_log(self):
        from PyQt6.QtWidgets import QFileDialog
        current = self.setup_page.roster_log_input.text().strip()
        start = str(Path(current).parent) if current else str(Path.home())
        path, _ = QFileDialog.getOpenFileName(
            self, "학교 전체 명단 파일 선택", start,
            "Excel 파일 (*.xlsx)"
        )
        if not path:
            return

        self.setup_page.roster_log_input.setText(path)

        # 열 매핑 팝업 자동 실행
        existing = load_app_config().get("roster_col_map", {})
        dlg = RosterColumnMapDialog(path, existing_map=existing, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self._roster_col_map = dlg.result_map
            # 설정 즉시 저장
            cfg = load_app_config()
            cfg["roster_log_path"] = path
            cfg["roster_col_map"] = self._roster_col_map
            save_app_config(cfg)
            QMessageBox.information(
                self, "열 매핑 저장",
                f"명단 파일 열 설정이 저장되었습니다.\n"
                f"학교명 열: {self._roster_col_map.get('col_school')}열"
            )

    def _on_email_sent_clicked(self):
        tab = self.main_page.main_tab
        tab.btn_email_hold.setChecked(False)
        tab.btn_email_sent.setChecked(True)

    def _on_email_hold_clicked(self):
        tab = self.main_page.main_tab
        tab.btn_email_sent.setChecked(False)
        tab.btn_email_hold.setChecked(True)

    def _on_record_roster(self):
        from PyQt6.QtWidgets import QMessageBox
        from datetime import date as _date

        if not self.roster_log_path:
            QMessageBox.warning(self, "명단 기록", "명단 파일이 설정되지 않았습니다.\n첫 화면에서 파일을 선택해 주세요.")
            return
        if not self.selected_school:
            QMessageBox.warning(self, "명단 기록", "학교가 선택되지 않았습니다.")
            return

        # 작업 내용 플래그 — 실행 결과 기준 (스캔 파일 존재 여부가 아니라 실제 처리 여부)
        run_result = self._last_run_result
        if run_result and run_result.ok and run_result.audit_summary:
            ic = run_result.audit_summary.get("input_counts", {})
            kind_flags = {
                "신입생": ic.get("freshmen", 0) > 0,
                "전입생": ic.get("transfer", 0) > 0,
                "전출생": ic.get("withdraw", 0) > 0,
                "교직원": ic.get("teacher", 0) > 0,
            }
        else:
            # 실행 결과가 없으면 스캔 결과로 fallback
            result = self._last_scan_result
            kind_flags = {
                "신입생": bool(result and getattr(result, "freshmen_file", None)),
                "전입생": bool(result and getattr(result, "transfer_file", None)),
                "전출생": bool(result and getattr(result, "withdraw_file", None)),
                "교직원": bool(result and getattr(result, "teacher_file", None)),
            }

        # 이메일 도착일자
        panel = self.main_page.status_panel
        email_arrived = None
        if panel.chk_email_arrived.isChecked():
            qd = panel.email_arrived_date.date()
            email_arrived = _date(qd.year(), qd.month(), qd.day())

        # 완료 이메일 발송일 (사이드바 기준)
        email_sent = None
        if panel.chk_email_sent.isChecked():
            qd = panel.email_sent_date.date()
            email_sent = _date(qd.year(), qd.month(), qd.day())

        ok1, msg1 = write_work_result(
            xlsx_path=self.roster_log_path,
            school_name=self.selected_school,
            worker=self.worker_name,
            kind_flags=kind_flags,
            email_arrived_date=email_arrived,
            col_map=self._roster_col_map,
        )
        if not ok1:
            QMessageBox.warning(self, "명단 기록 실패", msg1)
            return

        if panel.chk_email_sent.isChecked():
            ok2, msg2 = write_email_sent(
                xlsx_path=self.roster_log_path,
                school_name=self.selected_school,
                sent_date=email_sent,
                col_map=self._roster_col_map,
            )
            if not ok2:
                QMessageBox.warning(self, "명단 기록 경고", f"작업 내용은 기록됐지만 발송일 기록 중 오류:\n{msg2}")

        # 도착일을 마지막 세션으로 저장
        if panel.chk_email_arrived.isChecked():
            try:
                cfg = load_app_config()
                cfg["last_arrived_date"] = panel.email_arrived_date.date().toString("yyyy-MM-dd")
                save_app_config(cfg)
                self.app_config = cfg
            except Exception:
                pass

        self._pending_roster_log = False
        self.main_page.status_panel.btn_record_roster_sidebar.setEnabled(False)
        self.main_page.main_tab.email_log_status.setText(f"✓ 명단 기록 완료 — {self.selected_school}")

        # 작업 이력 저장 및 라벨 갱신
        entry = getattr(self, "_pending_history_entry", None)
        if entry and self.selected_school:
            from datetime import date as _date_cls
            school_year = getattr(self, "_school_year", _date_cls.today().year)
            try:
                save_work_history(school_year, self.selected_school, entry)
                lbl = self.main_page.status_panel.school_history_label
                count_str = " · ".join(f"{k} {v}명" for k, v in entry.get("counts", {}).items() if v)
                lbl.setText(
                    f"마지막 작업 · {entry['last_date']}"
                    + (f"\n{count_str}" if count_str else "")
                )
                lbl.show()
            except Exception:
                pass
            self._pending_history_entry = None

    def _check_pending_roster_before(self, action_label: str) -> bool:
        """
        pending 상태면 기록 여부를 물음.
        True 반환 = 계속 진행해도 됨 / False = 취소
        """
        if not self._pending_roster_log:
            return True
        from PyQt6.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            self,
            "명단 미기록",
            f"'{self.selected_school}' 작업 결과가 명단에 기록되지 않았습니다.\n"
            f"지금 기록하시겠어요?",
            QMessageBox.StandardButton.Yes |
            QMessageBox.StandardButton.No |
            QMessageBox.StandardButton.Cancel,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._on_record_roster()
            return True
        if reply == QMessageBox.StandardButton.Cancel:
            return False
        return True  # No → 기록 없이 진행

    def _load_school_list(self):
        school_input = self.main_page.status_panel.school_input
        school_list = self.main_page.status_panel.school_result_list

        school_input.blockSignals(True)
        school_input.clear()
        school_input.setPlaceholderText("학교명을 입력하세요")
        school_input.blockSignals(False)

        school_list.blockSignals(True)
        school_list.clear()
        school_list.blockSignals(False)

        self._school_names = []
        self._school_name_set = set()

        if not self.work_root:
            self.main_page.status_panel.school_status_label.setText("작업 폴더를 먼저 설정하세요.")
            return
        
        load_error = None

        try:
            names = load_all_school_names(self.roster_log_path, self._roster_col_map)
        except Exception as e:
            load_error = e
            names = []

        cleaned = []
        seen = set()
        for name in names or []:
            s = str(name).strip()
            if not s or s in seen:
                continue
            seen.add(s)
            cleaned.append(s)

        self._school_names = cleaned
        self._school_name_set = set(cleaned)

        school_list.blockSignals(True)
        school_list.clear()
        for name in self._school_names[:100]:
            school_list.addItem(QListWidgetItem(name))
        school_list.blockSignals(False)

        if load_error is not None:
            self.main_page.status_panel.school_status_label.setText(
                f"학교 목록 로드 실패: {type(load_error).__name__}"
            )
        elif self._school_names:
            self.main_page.status_panel.school_status_label.setText(
                f"학교명 검색 준비 완료 · {len(self._school_names)}개"
            )
        else:
            self.main_page.status_panel.school_status_label.setText(
                "학교 목록 0건: 명단 파일을 선택하고 열 매핑을 완료해 주세요."
            )

    def on_school_text_edited(self, text: str):
        self._pending_school_keyword = (text or "").strip()
        self.main_page.status_panel.school_result_list.show()
        self._school_search_timer.start()

    def _apply_school_search(self):
        school_list = self.main_page.status_panel.school_result_list
        keyword = self._pending_school_keyword

        if not self._school_names:
            self.main_page.status_panel.school_status_label.setText("학교 목록을 불러오지 못했습니다.")
            self.main_page.status_panel.btn_select_school.setEnabled(False)
            school_list.clear()
            return
        
        # 학교 목록 자체가 없으면 숨기기
        if not self._school_names:
            self.main_page.status_panel.school_status_label.setText("학교 목록을 불러오지 못했습니다.")
            self.main_page.status_panel.btn_select_school.setEnabled(False)
            school_list.clear()
            school_list.hide()
            return
        
        # 검색어 없으면 전체 목록 보여주고 리스트 보이기
        if not keyword:
            matched = self._school_names[:100]
            self.main_page.status_panel.school_status_label.setText(
                f"전체 학교 목록 {len(self._school_names)}개"
            )
            self.main_page.status_panel.btn_select_school.setEnabled(False)
            school_list.show()

        # 검색어 있는데 결과 있으면 보이기
        else:
            matched = [name for name in self._school_names if keyword.lower() in name.lower()][:100]
            if matched:
                self.main_page.status_panel.school_status_label.setText(f"검색 결과 {len(matched)}건")
                school_list.show()
            else:
                self.main_page.status_panel.school_status_label.setText("DB에 일치하는 학교가 없습니다.")
                school_list.hide()

        if not keyword:
            matched = self._school_names[:100]
            self.main_page.status_panel.school_status_label.setText(
                f"전체 학교 목록 {len(self._school_names)}개"
            )
            self.main_page.status_panel.btn_select_school.setEnabled(False)
        else:
            matched = [name for name in self._school_names if keyword.lower() in name.lower()][:100]
            if matched:
                self.main_page.status_panel.school_status_label.setText(f"검색 결과 {len(matched)}건")
            else:
                self.main_page.status_panel.school_status_label.setText("DB에 일치하는 학교가 없습니다.")

        school_list.blockSignals(True)
        school_list.clear()
        for name in matched:
            school_list.addItem(QListWidgetItem(name))
        school_list.blockSignals(False)

        self.validate_school_input(keyword)


    def validate_school_input(self, text: str):
        name = (text or "").strip()

        if not self.work_root or not name:
            self.main_page.status_panel.btn_select_school.setEnabled(False)
            if not name:
                self.main_page.status_panel.school_status_label.setText("학교명을 입력해 검색하세요.")
            return

        exact = name in self._school_name_set
        self.main_page.status_panel.btn_select_school.setEnabled(exact)

        if exact:
            self.main_page.status_panel.school_status_label.setText("적용 가능한 학교입니다.")
        else:
            matched = [school for school in self._school_names if name.lower() in school.lower()]
            if matched:
                self.main_page.status_panel.school_status_label.setText(f"검색 결과 {len(matched)}건")
            else:
                self.main_page.status_panel.school_status_label.setText("DB에 없는 학교입니다.")

    
    def on_school_item_clicked(self, item):
        if item is None:
            return
        text = item.text().strip()
        self.main_page.status_panel.school_input.setText(text)
        self.validate_school_input(text)

    def on_school_item_double_clicked(self, item):
        if item is None:
            return
        text = item.text().strip()
        self.main_page.status_panel.school_input.setText(text)
        self.validate_school_input(text)
        if text in self._school_name_set:
            self.apply_selected_school()



    def _refresh_notice_tab(self):

        templates = load_notice_templates(self.work_root)
        self._notice_templates = {}

        school = self.selected_school or ""
        year   = str(self.school_start_date.year) if self.school_start_date else ""
        prev_year = str(self.school_start_date.year - 1) if self.school_start_date else "" 
        month  = str(self.school_start_date.month) if self.school_start_date else ""
        day    = str(self.school_start_date.day) if self.school_start_date else ""
        domain = getattr(self, "_selected_domain", "")

        # #24 순서: 신규등록 > 반이동 > 교직원 > 2-6학년
        ORDER_KEYWORDS = ["신규등록", "반이동", "교직원", "2-6"]

        def _sort_key(k: str) -> tuple:
            k_lower = k.lower()
            for i, kw in enumerate(ORDER_KEYWORDS):
                if kw in k_lower:
                    return (i, k)
            return (len(ORDER_KEYWORDS), k)

        sorted_keys = sorted(templates.keys(), key=_sort_key)

        tab = self.main_page.main_tab
        tab.notice_list.blockSignals(True)
        tab.notice_list.clear()

        for key in sorted_keys:
            text = templates[key]
            replaced = (text
                .replace("{school_name}", school)
                .replace("{year}", year)
                .replace("{prev_year}", prev_year)
                .replace("{month}", month)
                .replace("{day}", day)
                .replace("{domain}", domain)
            )
            self._notice_templates[key] = replaced
            tab.notice_list.addItem(key)

        tab.notice_list.blockSignals(False)

        if self._notice_templates:
            tab.notice_list.setCurrentRow(0)
            first_key = list(self._notice_templates.keys())[0]
            tab.notice_text.setPlainText(self._notice_templates[first_key])
        else:
            tab.notice_text.setPlainText("notices/ 폴더에 .txt 파일이 없습니다.")

    def _on_notice_row_changed(self, row):
        tab = self.main_page.main_tab
        item = tab.notice_list.item(row)
        if item is None:
            return
        key = item.text()
        text = getattr(self, "_notice_templates", {}).get(key, "")
        tab.notice_text.setPlainText(text)

    def _on_copy_notice(self):
        QApplication.clipboard().setText(
            self.main_page.main_tab.notice_text.toPlainText()
        )

    def _on_reset_notice(self):
        self._refresh_notice_tab()

    def closeEvent(self, event):
        if not self._check_pending_roster_before("종료"):
            event.ignore()
            return
        event.accept()

    def _reset_scan_run_state(self):
        """학교 변경 또는 초기화 시 스캔/실행/안내문 탭 내용을 전부 비운다."""
        self._last_scan_result = None
        self._last_scan_logs = []
        self._last_run_logs = []
        self._current_output_files = []

        tab = self.main_page.main_tab

        # 스캔 탭
        tab.scan_message.setText("파일 내용 스캔을 실행해 주세요.")
        tab.scan_status_label.setText("스캔 전")
        tab.scan_status_label.setStyleSheet(_STATUS_STYLE_IDLE)
        tab.scan_preview_data = {}
        tab.current_preview_kind = None
        tab.preview_file_info.setText("파일: - | 시트: - | 헤더행: - | 시작행: -")
        tab.preview_warning_label.setText("학교를 선택하고 스캔을 실행해 주세요.")
        tab.preview_table.setRowCount(0)
        tab.preview_table.setColumnCount(0)
        tab.btn_toggle_viewer.setChecked(False)
        tab.viewer_body.setVisible(False)
        tab.btn_toggle_viewer.setText("펼치기 ▾")
        tab.btn_goto_run_tab.setEnabled(False)
        for kind, row in self._SCAN_TABLE_KIND_ROW.items():
            for c in range(1, 4):
                cell = tab.scan_table.item(row, c)
                if cell:
                    cell.setText("")
            spin_widget = tab.scan_table.cellWidget(row, 4)
            if spin_widget is not None:
                lbl = spin_widget.findChild(QLabel)
                if lbl:
                    lbl.setText("-")
                    lbl.setStyleSheet("font-weight: 700; color: #94A3B8; font-size: 12px;")
            chk_widget = tab.scan_table.cellWidget(row, 5)
            if chk_widget is not None:
                chk = chk_widget.findChild(QCheckBox)
                if chk:
                    chk.setChecked(False)

        # 학교 구분 UI
        tab.school_kind_warn_label.hide()
        tab.school_kind_row_widget.hide()
        self._school_kind_active = False
        self._school_kind_override = None

        # 스캔/실행 결과 버튼 초기화
        tab.scan_result_widget.setMaximumHeight(0)
        tab.scan_result_summary.setText("")
        tab.btn_goto_run_tab.setEnabled(False)
        tab.btn_goto_run_tab.setVisible(False)
        tab.run_result_widget.setMaximumHeight(0)
        tab.run_result_summary.setText("")
        tab.btn_goto_notice_tab.setEnabled(False)
        tab.btn_goto_notice_tab.setVisible(False)

        # 실행 탭
        tab.run_info.setText("먼저 스캔을 통과해야 실행할 수 있습니다.")
        tab.run_status_label.setText("실행 전")
        tab.run_status_label.setStyleSheet(_STATUS_STYLE_IDLE)
        tab.run_hold_warning.hide()
        tab.run_preview_data = {}
        tab.run_file_combo.blockSignals(True)
        tab.run_file_combo.clear()
        tab.run_file_combo.blockSignals(False)
        tab.run_file_list_widget.setRowCount(0)
        tab.run_file_list_widget.setFixedHeight(80)
        tab.run_sheet_tabs.clear()
        tab.run_preview_info.setText("시트: - | 행 수: -")
        tab.run_file_title.setText("실행 전")
        tab.export_path.setText("")
        for lbl in [tab.sum_school, tab.sum_year, tab.sum_freshmen, tab.sum_teacher,
                    tab.sum_transfer, tab.sum_withdraw, tab.sum_transfer_check, tab.sum_withdraw_check]:
            lbl.setText("-")

        self.main_page.step_bar.set_state(2, "idle")
        self.main_page.step_bar.set_state(3, "idle")
        self.main_page.step_bar.set_state(4, "idle")
        self.main_page.status_panel.update_grade_map(state="default")
        self._reset_diff_result()
        self.main_page.main_tab.tabs.setCurrentIndex(0)

        # 안내문 탭
        tab.btn_email_sent.setChecked(False)
        tab.btn_email_hold.setChecked(False)
        tab.btn_record_roster.setEnabled(False)
        tab.email_log_status.setText("")
        tab.btn_run_dup_only.setChecked(False)
        tab.notice_list.blockSignals(True)
        tab.notice_list.clear()
        tab.notice_list.blockSignals(False)
        tab.notice_text.setPlainText("")
        self._pending_roster_log = False
        self._last_run_result = None
        self.main_page.status_panel.chk_email_arrived.setChecked(False)

    def _on_new_school_clicked(self):
        from PyQt6.QtWidgets import QMessageBox
        if not self._check_pending_roster_before("새 학교 시작"):
            return
        warnings = []
        sp = self.main_page.status_panel
        if not sp.chk_email_arrived.isChecked():
            warnings.append("도착일 기록")
        if not sp.chk_email_sent.isChecked():
            warnings.append("발송일 기록")
        if self._pending_roster_log:
            warnings.append("명단 기록")

        if warnings:
            reply = QMessageBox.question(
                self, "새 학교 시작",
                f"아직 완료되지 않은 항목이 있습니다:\n· " + "\n· ".join(warnings) +
                "\n\n그래도 새 학교 작업을 시작하시겠습니까?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
        else:
            reply = QMessageBox.question(
                self, "새 학교 시작",
                "새 학교 작업을 시작하시겠습니까?\n현재 작업 내용이 초기화됩니다.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        self._do_reset_for_new_school()

    def _do_reset_for_new_school(self):
        """학교 초기화 공통 처리"""
        self._reset_scan_run_state()
        self.selected_school = None
        sp = self.main_page.status_panel
        sp.school_input.setText("")
        sp.school_status_label.setText("학교명을 입력해 검색하세요.")
        sp.school_status_label.show()
        sp.current_school.setText("-")
        sp.school_history_label.setText("")
        sp.school_history_label.hide()
        sp.last_work_label.setText("")
        sp.last_work_label.hide()
        sp.chk_email_arrived.setChecked(False)
        sp.chk_email_sent.setChecked(False)
        sp.btn_record_roster_sidebar.setEnabled(False)
        self._reset_school_button()
        self.main_page.main_tab.btn_scan.setEnabled(False)
        self.main_page.main_tab.btn_run.setEnabled(False)
        self.main_page.main_tab.btn_run_diff.setEnabled(False)
        self.main_page.header.school_name.setText("학교를 선택해 주세요")
        self.main_page.header.badge_status.setText("상태 대기")
        self.main_page.step_bar.set_state(1, "idle")
        self.main_page.step_bar.set_state(2, "idle")
        self.main_page.step_bar.set_state(3, "idle")
        self.main_page.step_bar.set_state(4, "idle")
        self.main_page.main_tab.tabs.setCurrentIndex(0)
        sp.school_input.setFocus()

    def _reset_school_button(self):
        btn = self.main_page.status_panel.btn_select_school
        btn.setText("적용")
        btn.setEnabled(False)
        btn.setStyleSheet("")

    def go_back_to_setup(self):
        from PyQt6.QtWidgets import QMessageBox
        if not self._check_pending_roster_before("초기화"):
            return
        reply = QMessageBox.question(
            self, "초기 설정으로 돌아가기",
            "초기 설정으로 돌아가시겠습니까?\n현재 작업 내용이 모두 초기화됩니다.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        self._do_reset_for_new_school()
        self.main_page.step_bar.set_state(0, "idle")
        self.stack.setCurrentIndex(0)

    def _reset_to_school_select(self):
        from PyQt6.QtWidgets import QMessageBox
        if not self._check_pending_roster_before("새 작업"):
            return
        reply = QMessageBox.question(
            self, "새 작업 시작",
            "새 작업을 시작하시겠습니까?\n현재 스캔/실행 결과가 초기화됩니다.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        self._do_reset_for_new_school()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet(APP_QSS)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())